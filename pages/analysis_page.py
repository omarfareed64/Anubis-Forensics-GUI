import csv
import os
import struct
import uuid
import codecs
import hashlib
import re
import time
import json
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFrame, QMessageBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QLineEdit, QComboBox, QGroupBox, QGridLayout,
    QStatusBar, QProgressBar, QFileDialog, QAction, QMenu, QApplication, QTextEdit,
    QListWidget, QListWidgetItem, QScrollArea, QTabWidget
)
from PyQt5.QtGui import QFont, QColor, QKeySequence
from PyQt5.QtCore import Qt, pyqtSignal, QThread, pyqtSignal as Signal, QUrl
from PyQt5.QtWebEngineWidgets import QWebEngineView

# Third-party imports for SRUM
try:
    import pyesedb
    import openpyxl
    from Registry import Registry
    SRUM_IMPORTS_AVAILABLE = True
except ImportError:
    SRUM_IMPORTS_AVAILABLE = False

from .base_page import BasePage, COLOR_ORANGE, COLOR_DARK, COLOR_GRAY, TAB_NAMES
from services.web_artifact_extractor import extract_all_web_artifacts
from services.usb_analyzer import get_usb_devices
from services.registry_analyzer import RegistryAnalyzer
from datetime import datetime, timedelta

class WebArtifactThread(QThread):
    """Worker thread for extracting web artifacts."""
    finished = Signal(dict)

    def __init__(self, params, parent=None):
        super().__init__(parent)
        self.params = params

    def run(self):
        """Execute the extraction script."""
        result = extract_all_web_artifacts(
            remote_ip=self.params.get('remote_ip'),
            domain=self.params.get('remote_domain'),
            username=self.params.get('remote_user'),
            password=self.params.get('remote_password')
        )
        self.finished.emit(result)

class UsbDeviceThread(QThread):
    """Worker thread for scanning local USB device history."""
    finished = Signal(list)

    def run(self):
        """Execute the USB device scan."""
        devices = get_usb_devices()
        self.finished.emit(devices)

class RegistryWorker(QThread):
    """Worker thread for registry operations"""
    progress_updated = pyqtSignal(str)
    operation_completed = pyqtSignal(str, bool, str)
    header_output = pyqtSignal(str)  # For header parsing output
    
    def __init__(self, analyzer, operation, **kwargs):
        super().__init__()
        self.analyzer = analyzer
        self.operation = operation
        self.kwargs = kwargs
        
        # Connect the analyzer's signals to our signals
        self.analyzer.progress_updated.connect(self.progress_updated.emit)
        self.analyzer.operation_completed.connect(self.operation_completed.emit)
        self.analyzer.header_output.connect(self.header_output.emit)
        
    def run(self):
        # This will call the appropriate method on the RegistryAnalyzer instance
        operation_func = getattr(self.analyzer, self.operation, None)
        if operation_func:
            # We need to unpack the kwargs dict to pass them as arguments
            operation_func(**self.kwargs)

# --- SRUM Analyzer Logic ---
if SRUM_IMPORTS_AVAILABLE:
    class SrumAnalyzer:
        """
        Encapsulates the logic for SRUM database analysis.
        This implementation is based on the more comprehensive reference script.
        """

        def __init__(self, srum_path, template_path, reg_hive_path=None):
            self.srum_path = srum_path
            self.template_path = template_path
            self.reg_hive_path = reg_hive_path
            self.template_tables = {}
            self.template_lookups = {}
            self.id_table = {}
            self.interface_table = {}
            self.regsids = {}

        def analyze(self):
            """Main method to run the analysis."""
            if self.reg_hive_path:
                self.interface_table = self._load_interfaces(self.reg_hive_path)
                self.regsids = self._load_registry_sids(self.reg_hive_path)

            try:
                ese_db = pyesedb.file()
                ese_db.open(self.srum_path)
            except Exception as e:
                raise IOError(f"Could not open the specified SRUM file: {e}")

            try:
                template_wb = openpyxl.load_workbook(filename=self.template_path)
            except Exception as e:
                ese_db.close()
                raise IOError(f"Could not open the specified template file: {e}")

            self.template_tables = self._load_template_tables(template_wb)
            self.template_lookups = self._load_template_lookups(template_wb)
            if self.regsids:
                self.template_lookups.setdefault("Known SIDS", {}).update(self.regsids)
            
            self.id_table = self._load_srumid_lookups(ese_db)

            skip_tables = ['MSysObjects', 'MSysObjectsShadow', 'MSysObjids', 'MSysLocales', 'SruDbIdMapTable']
            all_tables_data, message = self._process_srum_tables(ese_db, skip_tables)

            ese_db.close()
            return all_tables_data, message

        def _process_srum_tables(self, ese_db, skip_tables):
            all_tables_data = {}
            for table_num in range(ese_db.number_of_tables):
                ese_table = ese_db.get_table(table_num)
                if ese_table.name in skip_tables:
                    continue

                tname = self._ese_table_guid_to_name(ese_table)
                num_recs = self._ese_table_record_count(ese_table)
                if not num_recs:
                    continue

                table_data = []
                column_names = [x.name for x in ese_table.columns]
                
                header_row = []
                if ese_table.name in self.template_tables:
                    _, tfields = self.template_tables.get(ese_table.name)
                    for eachcol in ese_table.columns:
                        if eachcol.name in tfields:
                            _, _, cell_value = tfields.get(eachcol.name)
                            header_row.append(cell_value)
                        else:
                            header_row.append(eachcol.name)
                else:
                    header_row = [x.name for x in ese_table.columns]
                table_data.append(header_row)

                for row_num in range(num_recs):
                    ese_row = self._ese_table_get_record(ese_table, row_num)
                    if ese_row is None: continue
                    
                    gui_row = []
                    for col_num in range(ese_table.number_of_columns):
                        val = self._smart_retrieve(ese_table, row_num, col_num)
                        if val == "Error": val = f"WARNING: Invalid Column Name {column_names[col_num]}"
                        elif val is None: val = "None"
                        elif ese_table.name in self.template_tables:
                            _, tfields = self.template_tables.get(ese_table.name)
                            if column_names[col_num] in tfields:
                                _, cformat, _ = tfields.get(column_names[col_num])
                                val = self._format_output_for_gui(val, cformat)
                        gui_row.append(str(val))
                    table_data.append(gui_row)

                all_tables_data[tname] = table_data
            return all_tables_data, "Finished processing all tables."
        
        def _load_registry_sids(self, reg_file):
            sids = {}
            try:
                reg_handle = Registry.Registry(reg_file)
                profile_key = reg_handle.open(r"Microsoft\Windows NT\CurrentVersion\ProfileList")
                for eachsid in profile_key.subkeys():
                    sids_path = eachsid.value("ProfileImagePath").value()
                    sids[eachsid.name()] = sids_path.split("\\")[-1]
            except Exception: return {}
            return sids

        def _load_interfaces(self, reg_file):
            profile_lookup = {}
            try:
                reg_handle = Registry.Registry(reg_file)
                int_keys = reg_handle.open('Microsoft\\WlanSvc\\Interfaces')
                for eachinterface in int_keys.subkeys():
                    if not hasattr(eachinterface, "subkey") or "Profiles" not in [s.name() for s in eachinterface.subkeys()]: continue
                    for eachprofile in eachinterface.subkey("Profiles").subkeys():
                        profileid_val = [x for x in list(eachprofile.values()) if x.name() == "ProfileIndex"]
                        if not profileid_val: continue
                        profileid = profileid_val[0].value()
                        if "MetaData" not in [s.name() for s in eachprofile.subkeys()]: continue
                        metadata = list(eachprofile.subkey("MetaData").values())
                        for eachvalue in metadata:
                            if eachvalue.name() in ["Channel Hints", "Band Channel Hints"]:
                                channelhintraw, hintlength = eachvalue.value(), struct.unpack("I", eachvalue.value()[0:4])[0]
                                name = channelhintraw[4:hintlength + 4]
                                profile_lookup[str(profileid)] = name.decode(encoding="latin1")
            except Exception: pass
            return profile_lookup

        def _load_srumid_lookups(self, database):
            id_lookup = {}
            try:
                lookup_table = database.get_table_by_name('SruDbIdMapTable')
                column_lookup = {x.name: i for i, x in enumerate(lookup_table.columns)}
            except (IOError, AttributeError): return {}
            for rec_num in range(self._ese_table_record_count(lookup_table)):
                blob = self._smart_retrieve(lookup_table, rec_num, column_lookup['IdBlob'])
                id_type = self._smart_retrieve(lookup_table, rec_num, column_lookup['IdType'])
                if id_type == 3: blob = self._binary_sid_to_string_sid(blob)
                elif blob not in ["Empty", "Error"]: blob = self._blob_to_string(blob)
                id_lookup[self._smart_retrieve(lookup_table, rec_num, column_lookup['IdIndex'])] = blob
            return id_lookup

        def _load_template_lookups(self, wb):
            lookups = {}
            for name in wb.sheetnames:
                if name.lower().startswith("lookup-"):
                    lookup_name = name.split("-")[1]
                    sheet, table = wb[name], {}
                    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
                        if row[0] is not None: table[row[0]] = row[1]
                    lookups[lookup_name] = table
            return lookups

        def _load_template_tables(self, wb):
            tables = {}
            for name in wb.sheetnames:
                if name.lower().startswith("lookup-"): continue
                sheet = wb[name]
                ese_table = sheet.cell(row=1, column=1).value
                if not ese_table: continue
                fields = {}
                for col in range(1, sheet.max_column + 1):
                    field_name = sheet.cell(row=2, column=col).value
                    if not field_name: break
                    fields[field_name] = (
                        sheet.cell(row=4, column=col).style,
                        sheet.cell(row=3, column=col).value,
                        sheet.cell(row=4, column=col).value or field_name
                    )
                tables[ese_table] = (name, fields)
            return tables

        def _smart_retrieve(self, ese_table, ese_record_num, column_number):
            rec = self._ese_table_get_record(ese_table, ese_record_num)
            if not rec: return "Error"
            
            col_type = rec.get_column_type(column_number)
            col_data = rec.get_value_data(column_number)
            
            if col_data is None: return "Empty"

            try:
                if col_type == pyesedb.column_types.BINARY_DATA: return codecs.encode(col_data,"HEX").decode()
                elif col_type == pyesedb.column_types.BOOLEAN: return struct.unpack('?',col_data)[0]
                elif col_type == pyesedb.column_types.DATE_TIME: return self._ole_timestamp(col_data)
                elif col_type == pyesedb.column_types.DOUBLE_64BIT: return struct.unpack('d',col_data)[0]
                elif col_type == pyesedb.column_types.FLOAT_32BIT: return struct.unpack('f',col_data)[0]
                elif col_type == pyesedb.column_types.GUID: return str(uuid.UUID(bytes=col_data))
                elif col_type == pyesedb.column_types.INTEGER_16BIT_SIGNED: return struct.unpack('h',col_data)[0]
                elif col_type == pyesedb.column_types.INTEGER_16BIT_UNSIGNED: return struct.unpack('H',col_data)[0]
                elif col_type == pyesedb.column_types.INTEGER_32BIT_SIGNED: return struct.unpack('i',col_data)[0]
                elif col_type == pyesedb.column_types.INTEGER_32BIT_UNSIGNED: return struct.unpack('I',col_data)[0]
                elif col_type == pyesedb.column_types.INTEGER_64BIT_SIGNED: return struct.unpack('q',col_data)[0]
                elif col_type == pyesedb.column_types.INTEGER_8BIT_UNSIGNED: return struct.unpack('B',col_data)[0]
                elif col_type == pyesedb.column_types.LARGE_BINARY_DATA: return codecs.encode(col_data,"HEX").decode()
                elif col_type == pyesedb.column_types.LARGE_TEXT: return self._blob_to_string(col_data)
                elif col_type == pyesedb.column_types.SUPER_LARGE_VALUE: return codecs.encode(col_data,"HEX").decode()
                elif col_type == pyesedb.column_types.TEXT: return self._blob_to_string(col_data)
            except (struct.error, TypeError):
                return self._blob_to_string(col_data) # Fallback on error
            
            return self._blob_to_string(col_data)

        def _format_output_for_gui(self, val, fmt):
            if val is None: return "None"
            if fmt is None: return str(val)
            
            fmt_lower = fmt.lower()
            try:
                if fmt_lower.startswith("ole"):
                    if isinstance(val, datetime): val = val.strftime(fmt[4:] if ":" in fmt else '%Y-%m-%d %H:%M:%S')
                elif fmt_lower.startswith("file"):
                    ft = self._file_timestamp(val)
                    if isinstance(ft, datetime): val = ft.strftime(fmt[5:] if ":" in fmt else '%Y-%m-%d %H:%M:%S')
                    else: val = ft
                elif fmt_lower.startswith("lookup-"):
                    lookup_name = fmt.split("-")[1]
                    val = self.template_lookups.get(lookup_name, {}).get(val, val)
                elif fmt_lower == "lookup_id": val = self.id_table.get(val, f"Unknown ID ({val})")
                elif fmt_lower == "lookup_luid":
                    inttype = struct.unpack(">H6B", codecs.decode(format(val,'016x'),'hex'))[0]
                    val = self.template_lookups.get("LUID Interfaces",{}).get(inttype,"")
                elif fmt_lower == "seconds": val = str(timedelta(seconds=val or 0))
                elif fmt_lower == "md5": val = hashlib.md5(str(val).encode()).hexdigest()
                elif fmt_lower == "sha1": val = hashlib.sha1(str(val).encode()).hexdigest()
                elif fmt_lower == "sha256": val = hashlib.sha256(str(val).encode()).hexdigest()
                elif fmt_lower == "base16": val = hex(val) if isinstance(val, int) else format(val,"08x")
                elif fmt_lower == "base2": val = format(val,"032b") if isinstance(val, int) else int(str(val),2)
                elif fmt_lower == "interface_id" and self.reg_hive_path: val = self.interface_table.get(str(val),"")
            except Exception: pass
            return str(val)

        def _binary_sid_to_string_sid(self, sid_hex):
            if not sid_hex: return ""
            try:
                sid = codecs.decode(sid_hex, "hex")
                sid_str = f"S-{sid[0]}"
                sub_auth_count = sid[1]
                id_auth = struct.unpack(">Q", b'\x00\x00' + sid[2:8])[0]
                sid_str += f"-{id_auth}"
                for i in range(sub_auth_count):
                    sub_auth = struct.unpack("<L", sid[8 + i*4 : 12 + i*4])[0]
                    sid_str += f"-{sub_auth}"
                sid_name = self.template_lookups.get("Known SIDS", {}).get(sid_str, 'unknown')
                return f"{sid_str} ({sid_name})"
            except Exception: return "Invalid SID"

        def _blob_to_string(self, blob):
            try:
                if isinstance(blob, str): chrblob = codecs.decode(blob, "hex")
                else: chrblob = blob
                
                if re.match(b'^(?:[^\x00]\x00)+\x00\x00$', chrblob): return chrblob.decode("utf-16-le").strip("\x00")
                elif re.match(b'^(?:\x00[^\x00])+\x00\x00$', chrblob): return chrblob.decode("utf-16-be").strip("\x00")
                else: return chrblob.decode("latin1").strip("\x00")
            except Exception:
                return codecs.encode(blob, 'hex').decode() if isinstance(blob, bytes) else str(blob)

        def _ole_timestamp(self, blob):
            """Converts a hex encoded OLE time stamp to a time string"""
            try:
                td, ts = str(struct.unpack("<d", blob)[0]).split(".")
                dt = datetime(1899, 12, 30, 0, 0, 0) + timedelta(days=int(td), seconds=86400 * float(f"0.{ts}"))
                return dt
            except Exception:
                return "Invalid OLE Timestamp"

        def _file_timestamp(self, n):
            try: return datetime(1601, 1, 1) + timedelta(microseconds=n / 10)
            except Exception: return "Invalid File Timestamp"
            
        def _ese_table_guid_to_name(self, table):
            return self.template_tables.get(table.name, (table.name,))[0]
            
        def _ese_table_record_count(self, ese_table):
            try: return ese_table.number_of_records
            except Exception: return 0

        def _ese_table_get_record(self, ese_table, row_num):
            try: return ese_table.get_record(row_num)
            except Exception: return None

    class SrumAnalysisThread(QThread):
        """Worker thread for running SRUM analysis."""
        finished = Signal(dict)
        def __init__(self, srum_path, template_path, reg_hive_path=None, parent=None):
            super().__init__(parent)
            self.srum_path = srum_path
            self.template_path = template_path
            self.reg_hive_path = reg_hive_path
        def run(self):
            try:
                analyzer = SrumAnalyzer(self.srum_path, self.template_path, self.reg_hive_path)
                all_tables, message = analyzer.analyze()
                self.finished.emit({"status": "success", "data": all_tables, "message": message})
            except Exception as e:
                self.finished.emit({"status": "error", "message": str(e)})

class AnalysisPage(BasePage):
    def __init__(self):
        super().__init__()
        self.connection_params = None
        self.selected_case_path = None
        self.web_artifact_thread = None
        self.usb_device_thread = None
        self.registry_worker_thread = None
        self.registry_analyzer = RegistryAnalyzer() # Add analyzer instance
        self.srum_analysis_thread = None
        self.usb_devices = [] # To store full list of devices
        self.displayed_usb_devices = [] # To store the currently visible list
        self._setup_ui()

    def _switch_right_panel_view(self, view_to_show):
        """Show the selected view and hide others."""
        self.web_view.setVisible(self.web_view == view_to_show)
        self.usb_view_container.setVisible(self.usb_view_container == view_to_show)
import csv
import os
<<<<<<< HEAD
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFrame, QMessageBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QLineEdit, QComboBox, QGroupBox, QGridLayout,
    QStatusBar, QProgressBar, QFileDialog, QAction, QMenu, QApplication, QTextEdit,
    QListWidget, QListWidgetItem, QScrollArea
=======
import struct
import uuid
import codecs
import hashlib
import re
import time
import json
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFrame, QMessageBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QLineEdit, QComboBox, QGroupBox, QGridLayout,
    QStatusBar, QProgressBar, QFileDialog, QAction, QMenu, QApplication, QTabWidget, QTextEdit
>>>>>>> 99e81c0b01190573585baf4d5569a60786b8c6be
)
from PyQt5.QtGui import QFont, QColor, QKeySequence
from PyQt5.QtCore import Qt, pyqtSignal, QThread, pyqtSignal as Signal, QUrl
from PyQt5.QtWebEngineWidgets import QWebEngineView

# Third-party imports for SRUM
try:
    import pyesedb
    import openpyxl
    from Registry import Registry
    SRUM_IMPORTS_AVAILABLE = True
except ImportError:
    SRUM_IMPORTS_AVAILABLE = False

from .base_page import BasePage, COLOR_ORANGE, COLOR_DARK, COLOR_GRAY, TAB_NAMES
from services.web_artifact_extractor import extract_all_web_artifacts
from services.usb_analyzer import get_usb_devices
from services.registry_analyzer import RegistryAnalyzer
from datetime import datetime, timedelta

class WebArtifactThread(QThread):
    """Worker thread for extracting web artifacts."""
    finished = Signal(dict)

    def __init__(self, params, parent=None):
        super().__init__(parent)
        self.params = params

    def run(self):
        """Execute the extraction script."""
        result = extract_all_web_artifacts(
            remote_ip=self.params.get('remote_ip'),
            domain=self.params.get('remote_domain'),
            username=self.params.get('remote_user'),
            password=self.params.get('remote_password')
        )
        self.finished.emit(result)

class UsbDeviceThread(QThread):
    """Worker thread for scanning local USB device history."""
    finished = Signal(list)

    def run(self):
        """Execute the USB device scan."""
        devices = get_usb_devices()
        self.finished.emit(devices)

<<<<<<< HEAD
class RegistryWorker(QThread):
    """Worker thread for registry operations"""
    progress_updated = pyqtSignal(str)
    operation_completed = pyqtSignal(str, bool, str)
    header_output = pyqtSignal(str)  # For header parsing output
    
    def __init__(self, analyzer, operation, **kwargs):
        super().__init__()
        self.analyzer = analyzer
        self.operation = operation
        self.kwargs = kwargs
        
        # Connect the analyzer's signals to our signals
        self.analyzer.progress_updated.connect(self.progress_updated.emit)
        self.analyzer.operation_completed.connect(self.operation_completed.emit)
        self.analyzer.header_output.connect(self.header_output.emit)
        
    def run(self):
        # This will call the appropriate method on the RegistryAnalyzer instance
        operation_func = getattr(self.analyzer, self.operation, None)
        if operation_func:
            # We need to unpack the kwargs dict to pass them as arguments
            success, message = operation_func(**self.kwargs)
            self.operation_completed.emit(self.operation, success, message)
=======
# --- SRUM Analyzer Logic ---
# Note: This large class is included here to avoid file creation issues.
# It is recommended to move this to its own file in `services/`.

if SRUM_IMPORTS_AVAILABLE:
    class SrumAnalyzer:
        """
        Encapsulates the logic for SRUM database analysis.
        This implementation is based on the more comprehensive reference script.
        """

        def __init__(self, srum_path, template_path, reg_hive_path=None):
            self.srum_path = srum_path
            self.template_path = template_path
            self.reg_hive_path = reg_hive_path
            self.template_tables = {}
            self.template_lookups = {}
            self.id_table = {}
            self.interface_table = {}
            self.regsids = {}

        def analyze(self):
            """Main method to run the analysis."""
            if self.reg_hive_path:
                self.interface_table = self._load_interfaces(self.reg_hive_path)
                self.regsids = self._load_registry_sids(self.reg_hive_path)

            try:
                ese_db = pyesedb.file()
                ese_db.open(self.srum_path)
            except Exception as e:
                raise IOError(f"Could not open the specified SRUM file: {e}")

            try:
                template_wb = openpyxl.load_workbook(filename=self.template_path)
            except Exception as e:
                ese_db.close()
                raise IOError(f"Could not open the specified template file: {e}")

            self.template_tables = self._load_template_tables(template_wb)
            self.template_lookups = self._load_template_lookups(template_wb)
            if self.regsids:
                self.template_lookups.setdefault("Known SIDS", {}).update(self.regsids)
            
            self.id_table = self._load_srumid_lookups(ese_db)

            skip_tables = ['MSysObjects', 'MSysObjectsShadow', 'MSysObjids', 'MSysLocales', 'SruDbIdMapTable']
            all_tables_data, message = self._process_srum_tables(ese_db, skip_tables)

            ese_db.close()
            return all_tables_data, message

        def _process_srum_tables(self, ese_db, skip_tables):
            all_tables_data = {}
            for table_num in range(ese_db.number_of_tables):
                ese_table = ese_db.get_table(table_num)
                if ese_table.name in skip_tables:
                    continue

                tname = self._ese_table_guid_to_name(ese_table)
                num_recs = self._ese_table_record_count(ese_table)
                if not num_recs:
                    continue

                table_data = []
                column_names = [x.name for x in ese_table.columns]
                
                header_row = []
                if ese_table.name in self.template_tables:
                    _, tfields = self.template_tables.get(ese_table.name)
                    for eachcol in ese_table.columns:
                        if eachcol.name in tfields:
                            _, _, cell_value = tfields.get(eachcol.name)
                            header_row.append(cell_value)
                        else:
                            header_row.append(eachcol.name)
                else:
                    header_row = [x.name for x in ese_table.columns]
                table_data.append(header_row)

                for row_num in range(num_recs):
                    ese_row = self._ese_table_get_record(ese_table, row_num)
                    if ese_row is None: continue
                    
                    gui_row = []
                    for col_num in range(ese_table.number_of_columns):
                        val = self._smart_retrieve(ese_table, row_num, col_num)
                        if val == "Error": val = f"WARNING: Invalid Column Name {column_names[col_num]}"
                        elif val is None: val = "None"
                        elif ese_table.name in self.template_tables:
                            _, tfields = self.template_tables.get(ese_table.name)
                            if column_names[col_num] in tfields:
                                _, cformat, _ = tfields.get(column_names[col_num])
                                val = self._format_output_for_gui(val, cformat)
                        gui_row.append(str(val))
                    table_data.append(gui_row)

                all_tables_data[tname] = table_data
            return all_tables_data, "Finished processing all tables."
        
        def _load_registry_sids(self, reg_file):
            sids = {}
            try:
                reg_handle = Registry.Registry(reg_file)
                profile_key = reg_handle.open(r"Microsoft\Windows NT\CurrentVersion\ProfileList")
                for eachsid in profile_key.subkeys():
                    sids_path = eachsid.value("ProfileImagePath").value()
                    sids[eachsid.name()] = sids_path.split("\\")[-1]
            except Exception: return {}
            return sids

        def _load_interfaces(self, reg_file):
            profile_lookup = {}
            try:
                reg_handle = Registry.Registry(reg_file)
                int_keys = reg_handle.open('Microsoft\\WlanSvc\\Interfaces')
                for eachinterface in int_keys.subkeys():
                    if not hasattr(eachinterface, "subkey") or "Profiles" not in [s.name() for s in eachinterface.subkeys()]: continue
                    for eachprofile in eachinterface.subkey("Profiles").subkeys():
                        profileid_val = [x for x in list(eachprofile.values()) if x.name() == "ProfileIndex"]
                        if not profileid_val: continue
                        profileid = profileid_val[0].value()
                        if "MetaData" not in [s.name() for s in eachprofile.subkeys()]: continue
                        metadata = list(eachprofile.subkey("MetaData").values())
                        for eachvalue in metadata:
                            if eachvalue.name() in ["Channel Hints", "Band Channel Hints"]:
                                channelhintraw, hintlength = eachvalue.value(), struct.unpack("I", eachvalue.value()[0:4])[0]
                                name = channelhintraw[4:hintlength + 4]
                                profile_lookup[str(profileid)] = name.decode(encoding="latin1")
            except Exception: pass
            return profile_lookup

        def _load_srumid_lookups(self, database):
            id_lookup = {}
            try:
                lookup_table = database.get_table_by_name('SruDbIdMapTable')
                column_lookup = {x.name: i for i, x in enumerate(lookup_table.columns)}
            except (IOError, AttributeError): return {}
            for rec_num in range(self._ese_table_record_count(lookup_table)):
                blob = self._smart_retrieve(lookup_table, rec_num, column_lookup['IdBlob'])
                id_type = self._smart_retrieve(lookup_table, rec_num, column_lookup['IdType'])
                if id_type == 3: blob = self._binary_sid_to_string_sid(blob)
                elif blob not in ["Empty", "Error"]: blob = self._blob_to_string(blob)
                id_lookup[self._smart_retrieve(lookup_table, rec_num, column_lookup['IdIndex'])] = blob
            return id_lookup

        def _load_template_lookups(self, wb):
            lookups = {}
            for name in wb.sheetnames:
                if name.lower().startswith("lookup-"):
                    lookup_name = name.split("-")[1]
                    sheet, table = wb[name], {}
                    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
                        if row[0] is not None: table[row[0]] = row[1]
                    lookups[lookup_name] = table
            return lookups

        def _load_template_tables(self, wb):
            tables = {}
            for name in wb.sheetnames:
                if name.lower().startswith("lookup-"): continue
                sheet = wb[name]
                ese_table = sheet.cell(row=1, column=1).value
                if not ese_table: continue
                fields = {}
                for col in range(1, sheet.max_column + 1):
                    field_name = sheet.cell(row=2, column=col).value
                    if not field_name: break
                    fields[field_name] = (
                        sheet.cell(row=4, column=col).style,
                        sheet.cell(row=3, column=col).value,
                        sheet.cell(row=4, column=col).value or field_name
                    )
                tables[ese_table] = (name, fields)
            return tables

        def _smart_retrieve(self, ese_table, ese_record_num, column_number):
            rec = self._ese_table_get_record(ese_table, ese_record_num)
            if not rec: return "Error"
            
            col_type = rec.get_column_type(column_number)
            col_data = rec.get_value_data(column_number)
            
            if col_data is None: return "Empty"

            try:
                if col_type == pyesedb.column_types.BINARY_DATA: return codecs.encode(col_data,"HEX").decode()
                elif col_type == pyesedb.column_types.BOOLEAN: return struct.unpack('?',col_data)[0]
                elif col_type == pyesedb.column_types.DATE_TIME: return self._ole_timestamp(col_data)
                elif col_type == pyesedb.column_types.DOUBLE_64BIT: return struct.unpack('d',col_data)[0]
                elif col_type == pyesedb.column_types.FLOAT_32BIT: return struct.unpack('f',col_data)[0]
                elif col_type == pyesedb.column_types.GUID: return str(uuid.UUID(bytes=col_data))
                elif col_type == pyesedb.column_types.INTEGER_16BIT_SIGNED: return struct.unpack('h',col_data)[0]
                elif col_type == pyesedb.column_types.INTEGER_16BIT_UNSIGNED: return struct.unpack('H',col_data)[0]
                elif col_type == pyesedb.column_types.INTEGER_32BIT_SIGNED: return struct.unpack('i',col_data)[0]
                elif col_type == pyesedb.column_types.INTEGER_32BIT_UNSIGNED: return struct.unpack('I',col_data)[0]
                elif col_type == pyesedb.column_types.INTEGER_64BIT_SIGNED: return struct.unpack('q',col_data)[0]
                elif col_type == pyesedb.column_types.INTEGER_8BIT_UNSIGNED: return struct.unpack('B',col_data)[0]
                elif col_type == pyesedb.column_types.LARGE_BINARY_DATA: return codecs.encode(col_data,"HEX").decode()
                elif col_type == pyesedb.column_types.LARGE_TEXT: return self._blob_to_string(col_data)
                elif col_type == pyesedb.column_types.SUPER_LARGE_VALUE: return codecs.encode(col_data,"HEX").decode()
                elif col_type == pyesedb.column_types.TEXT: return self._blob_to_string(col_data)
            except (struct.error, TypeError):
                return self._blob_to_string(col_data) # Fallback on error
            
            return self._blob_to_string(col_data)

        def _format_output_for_gui(self, val, fmt):
            if val is None: return "None"
            if fmt is None: return str(val)
            
            fmt_lower = fmt.lower()
            try:
                if fmt_lower.startswith("ole"):
                    if isinstance(val, datetime): val = val.strftime(fmt[4:] if ":" in fmt else '%Y-%m-%d %H:%M:%S')
                elif fmt_lower.startswith("file"):
                    ft = self._file_timestamp(val)
                    if isinstance(ft, datetime): val = ft.strftime(fmt[5:] if ":" in fmt else '%Y-%m-%d %H:%M:%S')
                    else: val = ft
                elif fmt_lower.startswith("lookup-"):
                    lookup_name = fmt.split("-")[1]
                    val = self.template_lookups.get(lookup_name, {}).get(val, val)
                elif fmt_lower == "lookup_id": val = self.id_table.get(val, f"Unknown ID ({val})")
                elif fmt_lower == "lookup_luid":
                    inttype = struct.unpack(">H6B", codecs.decode(format(val,'016x'),'hex'))[0]
                    val = self.template_lookups.get("LUID Interfaces",{}).get(inttype,"")
                elif fmt_lower == "seconds": val = str(timedelta(seconds=val or 0))
                elif fmt_lower == "md5": val = hashlib.md5(str(val).encode()).hexdigest()
                elif fmt_lower == "sha1": val = hashlib.sha1(str(val).encode()).hexdigest()
                elif fmt_lower == "sha256": val = hashlib.sha256(str(val).encode()).hexdigest()
                elif fmt_lower == "base16": val = hex(val) if isinstance(val, int) else format(val,"08x")
                elif fmt_lower == "base2": val = format(val,"032b") if isinstance(val, int) else int(str(val),2)
                elif fmt_lower == "interface_id" and self.reg_hive_path: val = self.interface_table.get(str(val),"")
            except Exception: pass
            return str(val)

        def _binary_sid_to_string_sid(self, sid_hex):
            if not sid_hex: return ""
            try:
                sid = codecs.decode(sid_hex, "hex")
                sid_str = f"S-{sid[0]}"
                sub_auth_count = sid[1]
                id_auth = struct.unpack(">Q", b'\x00\x00' + sid[2:8])[0]
                sid_str += f"-{id_auth}"
                for i in range(sub_auth_count):
                    sub_auth = struct.unpack("<L", sid[8 + i*4 : 12 + i*4])[0]
                    sid_str += f"-{sub_auth}"
                sid_name = self.template_lookups.get("Known SIDS", {}).get(sid_str, 'unknown')
                return f"{sid_str} ({sid_name})"
            except Exception: return "Invalid SID"

        def _blob_to_string(self, blob):
            try:
                if isinstance(blob, str): chrblob = codecs.decode(blob, "hex")
                else: chrblob = blob
                
                if re.match(b'^(?:[^\x00]\x00)+\x00\x00$', chrblob): return chrblob.decode("utf-16-le").strip("\x00")
                elif re.match(b'^(?:\x00[^\x00])+\x00\x00$', chrblob): return chrblob.decode("utf-16-be").strip("\x00")
                else: return chrblob.decode("latin1").strip("\x00")
            except Exception:
                return codecs.encode(blob, 'hex').decode() if isinstance(blob, bytes) else str(blob)

        def _ole_timestamp(self, blob):
            """Converts a hex encoded OLE time stamp to a time string"""
            try:
                td, ts = str(struct.unpack("<d", blob)[0]).split(".")
                dt = datetime(1899, 12, 30, 0, 0, 0) + timedelta(days=int(td), seconds=86400 * float(f"0.{ts}"))
                return dt
            except Exception:
                return "Invalid OLE Timestamp"

        def _file_timestamp(self, n):
            try: return datetime(1601, 1, 1) + timedelta(microseconds=n / 10)
            except Exception: return "Invalid File Timestamp"
            
        def _ese_table_guid_to_name(self, table):
            return self.template_tables.get(table.name, (table.name,))[0]
            
        def _ese_table_record_count(self, ese_table):
            try: return ese_table.number_of_records
            except Exception: return 0

        def _ese_table_get_record(self, ese_table, row_num):
            try: return ese_table.get_record(row_num)
            except Exception: return None

    class SrumAnalysisThread(QThread):
        """Worker thread for running SRUM analysis."""
        finished = Signal(dict)

        def __init__(self, params, parent=None):
            super().__init__(parent)
            self.params = params

        def run(self):
            """Execute the analysis."""
            try:
                analyzer = SrumAnalyzer(
                    srum_path=self.params['srum_path'],
                    template_path=self.params['template_path'],
                    reg_hive_path=self.params.get('reg_path')
                )
                data, message = analyzer.analyze()
                self.finished.emit({"status": "success", "data": data, "message": message})
            except Exception as e:
                self.finished.emit({"status": "error", "message": str(e)})
>>>>>>> 99e81c0b01190573585baf4d5569a60786b8c6be

class AnalysisPage(BasePage):
    back_requested = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.connection_params = None
        self.web_artifact_thread = None
        self.usb_device_thread = None
<<<<<<< HEAD
        self.registry_worker_thread = None
        self.registry_analyzer = RegistryAnalyzer() # Add analyzer instance
=======
        self.srum_analysis_thread = None
>>>>>>> 99e81c0b01190573585baf4d5569a60786b8c6be
        self.usb_devices = [] # To store full list of devices
        self.displayed_usb_devices = [] # To store the currently visible list
        self.selected_case_path = None
        self.setup_page_content()
        self._select_tab_programmatically("Analyze Evidence")

    def set_connection_params(self, params):
        """Receive and store connection parameters."""
        self.connection_params = params

    def set_case_path(self, case_path):
        self.selected_case_path = case_path
        if case_path:
            base_output = os.path.join(self.selected_case_path, "registry_analysis")
            self.acquire_output_dir_input.setText(os.path.join(base_output, "acquired_hives"))
            self.analyze_input_dir.setText(os.path.join(base_output, "acquired_hives"))
            self.compare_output_dir.setText(os.path.join(base_output, "comparison_results"))
            self.logs_output_dir.setText(os.path.join(base_output, "recovered_hives"))

    def _switch_right_panel_view(self, view_to_show):
        """Manages visibility of widgets in the right panel."""
        self.web_view.setVisible(self.web_view == view_to_show)
        self.usb_view_container.setVisible(self.usb_view_container == view_to_show)
<<<<<<< HEAD
        self.registry_view_container.setVisible(self.registry_view_container == view_to_show)
=======
        self.srum_tab_widget.setVisible(self.srum_tab_widget == view_to_show)
        self.memory_view_container.setVisible(self.memory_view_container == view_to_show)
>>>>>>> 99e81c0b01190573585baf4d5569a60786b8c6be
        self.placeholder_label.setVisible(self.placeholder_label == view_to_show)

    def setup_page_content(self):
        """Setup the page-specific content for the analysis page"""
        # Add tab bar
        self.main_layout.addLayout(self._setup_tab_bar(TAB_NAMES))
        self.main_layout.addSpacing(40)

        # Main content layout
        content_layout = QHBoxLayout()
        content_layout.setSpacing(20)
        content_layout.setContentsMargins(40, 0, 40, 20)

        # Left side (buttons)
        left_panel = QFrame()
        left_panel.setFixedWidth(250)
        left_panel.setStyleSheet("background-color: white; border-radius: 18px;")
        
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(20, 20, 20, 20)
        left_layout.setSpacing(15)
        left_layout.setAlignment(Qt.AlignTop)

        artifact_buttons = ["MEMORY", "WEB", "SRUM", "REGISTRY", "USB"]
        for artifact_name in artifact_buttons:
            button = QPushButton(artifact_name)
            button.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
            button.setStyleSheet(f"""
                QPushButton {{
                    background-color: {COLOR_DARK};
                    color: white;
                    border: none;
                    border-radius: 8px;
                    padding: 12px;
                    text-align: center;
                }}
                QPushButton:hover {{
                    background-color: {COLOR_ORANGE};
                }}
            """)
            button.clicked.connect(lambda ch, name=artifact_name: self.on_artifact_button_click(name))
            left_layout.addWidget(button)
        
        content_layout.addWidget(left_panel)

        # Right side (display area)
        self.right_panel = QFrame()
        self.right_panel.setStyleSheet(f"background-color: white; border: 2px solid {COLOR_DARK}; border-radius: 18px;")
        
        right_layout = QVBoxLayout(self.right_panel)
        right_layout.setContentsMargins(2, 2, 2, 2)

        # Web view for web artifacts
        self.web_view = QWebEngineView()
        right_layout.addWidget(self.web_view)

        # --- SRUM View Container ---
        self.srum_tab_widget = QTabWidget()
        self.srum_tab_widget.setStyleSheet("""
            QTabBar::tab {
                background: #f0f0f0;
                border: 1px solid #ccc;
                border-bottom-color: #c2c7d5;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                min-width: 8ex;
                padding: 8px;
                font-family: 'Segoe UI';
                font-weight: bold;
            }
            QTabBar::tab:selected {
                background: white;
                border-color: #9B9B9B;
                border-bottom-color: white;
            }
            QTabWidget::pane {
                border: 1px solid #ccc;
                border-top: none;
            }
        """)
        right_layout.addWidget(self.srum_tab_widget)
        
        # --- USB View Container ---
        self.usb_view_container = QWidget()
        self.usb_view_container.setStyleSheet("""
            QGroupBox {
                border: 1px solid #ccc;
                border-radius: 6px;
                margin-top: 10px;
                font-weight: bold;
                background-color: #f7f7f7;
                font-family: 'Segoe UI';
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 5px;
                font-family: 'Segoe UI';
            }
            QPushButton#exportButton { 
                background-color: #17a2b8; 
                font-family: 'Segoe UI';
            }
            QPushButton#exportButton:hover { background-color: #138496; }
            QPushButton#forensicButton { 
                background-color: #dc3545; 
                font-family: 'Segoe UI';
            }
            QPushButton#forensicButton:hover { background-color: #c82333; }
            QProgressBar {
                font-family: 'Segoe UI';
            }
        """)
        usb_layout = QVBoxLayout(self.usb_view_container)
        usb_layout.setContentsMargins(0, 0, 0, 0)
        usb_layout.setSpacing(10)

        # Controls Panel
        control_panel = QGroupBox("Controls")
        control_panel.setFont(QFont("Segoe UI", 9))
        control_layout = QGridLayout(control_panel)
        control_layout.setContentsMargins(15, 25, 15, 15)
        control_layout.setSpacing(10)
        
        self.usb_search_box = QLineEdit()
        self.usb_search_box.setPlaceholderText("Type to filter devices...")
        self.usb_search_box.setClearButtonEnabled(True)

        self.usb_search_box.setFont(QFont("Segoe UI", 9))
        self.usb_search_box.textChanged.connect(self.apply_usb_filters)

        self.usb_time_filter = QComboBox()
        self.usb_time_filter.addItems(["All Time", "Last 7 Days", "Last 30 Days", "Last 90 Days", "Last Year"])
        self.usb_time_filter.setFont(QFont("Segoe UI", 9))
        self.usb_time_filter.currentIndexChanged.connect(self.apply_usb_filters)
        
        self.export_button = QPushButton("Export to CSV")
        self.export_button.setObjectName("exportButton")
        self.export_button.setFont(QFont("Segoe UI", 9))
        self.export_button.clicked.connect(self.export_usb_csv)
        
        self.forensic_button = QPushButton("Forensic Analysis")
        self.forensic_button.setObjectName("forensicButton")
        self.forensic_button.setFont(QFont("Segoe UI", 9))
        self.forensic_button.clicked.connect(self.perform_forensic_analysis)

        search_label = QLabel("Search")
        search_label.setFont(QFont("Segoe UI", 9))
        time_label = QLabel("Time Range:")
        time_label.setFont(QFont("Segoe UI", 9))
        
        control_layout.addWidget(search_label, 0, 0)
        control_layout.addWidget(self.usb_search_box, 0, 1)
        control_layout.addWidget(time_label, 0, 2)
        control_layout.addWidget(self.usb_time_filter, 0, 3)
        control_layout.addWidget(self.export_button, 1, 0, 1, 2)
        control_layout.addWidget(self.forensic_button, 1, 2, 1, 2)
        control_layout.setColumnStretch(1, 1)

        usb_layout.addWidget(control_panel)

        # USB Table
        self.usb_table_view = QTableWidget()
        self.usb_table_view.setSortingEnabled(True)
        self.usb_table_view.setAlternatingRowColors(True)
        self.usb_table_view.setSelectionBehavior(QTableWidget.SelectRows)
        self.usb_table_view.setEditTriggers(QTableWidget.NoEditTriggers)
        self.usb_table_view.verticalHeader().setVisible(False)
        self.usb_table_view.horizontalHeader().setStretchLastSection(True)
        self.usb_table_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.usb_table_view.customContextMenuRequested.connect(self.show_usb_context_menu)
        self.usb_table_view.doubleClicked.connect(self.show_usb_device_details)
        self.usb_table_view.setFont(QFont("Segoe UI", 9))
        self.usb_table_view.horizontalHeader().setStyleSheet(f"""
            QHeaderView::section {{
                background-color: {COLOR_DARK};
                color: white;
                padding: 6px;
                font-weight: bold;
                font-family: 'Segoe UI';
            }}
        """)
        usb_layout.addWidget(self.usb_table_view, 1)

        # Status Bar
        self.usb_status_bar = QStatusBar()
        self.usb_status_bar.setSizeGripEnabled(False)
        self.usb_status_bar.setFont(QFont("Segoe UI", 9))
        self.usb_device_count_label = QLabel()
        self.usb_device_count_label.setFont(QFont("Segoe UI", 9))
        self.usb_status_bar.addPermanentWidget(self.usb_device_count_label)
        self.usb_progress_bar = QProgressBar()
        self.usb_progress_bar.setMaximumWidth(250)
        self.usb_progress_bar.setVisible(False)
        self.usb_progress_bar.setFont(QFont("Segoe UI", 9))
        self.usb_status_bar.addPermanentWidget(self.usb_progress_bar)
        usb_layout.addWidget(self.usb_status_bar)

        right_layout.addWidget(self.usb_view_container)

<<<<<<< HEAD
        # --- Registry View Container ---
        self.registry_view_container = self.create_registry_view()
        right_layout.addWidget(self.registry_view_container)
=======
        # --- Memory Analysis View Container ---
        self.memory_view_container = QWidget()
        self._setup_memory_analysis_view()
        right_layout.addWidget(self.memory_view_container)
>>>>>>> 99e81c0b01190573585baf4d5569a60786b8c6be

        # Placeholder label for messages
        self.placeholder_label = QLabel("Select an artifact to view details")
        self.placeholder_label.setFont(QFont("Segoe UI", 16))
        self.placeholder_label.setAlignment(Qt.AlignCenter)
        self.placeholder_label.setStyleSheet("color: #aaa;")
        right_layout.addWidget(self.placeholder_label)

        content_layout.addWidget(self.right_panel, 1) # Stretch factor of 1

        self.main_layout.addLayout(content_layout, 1)

        self._switch_right_panel_view(self.placeholder_label) # Show placeholder initially

    def _setup_memory_analysis_view(self):
        """Creates the entire UI for the memory analysis section."""
        memory_layout = QVBoxLayout(self.memory_view_container)
        memory_layout.setContentsMargins(10, 10, 10, 10)
        memory_layout.setSpacing(10)

        # --- Top Tab Bar ---
        tabs_layout = QHBoxLayout()
        self.memory_tabs = {}
        memory_tab_names = ["Core Analysis Files", "Volatility", "Memory Dumps"]
        for name in memory_tab_names:
            button = QPushButton(name)
            button.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
            button.setCheckable(True)
            button.clicked.connect(self._on_memory_tab_click)
            tabs_layout.addWidget(button)
            self.memory_tabs[name] = button
        
        memory_layout.addLayout(tabs_layout)

        # --- Main Content Area ---
        content_layout = QHBoxLayout()
        
        # Left Panel (Sub-options)
        self.memory_left_panel = QFrame()
        self.memory_left_panel.setFixedWidth(250)
        self.memory_left_panel.setStyleSheet("background-color: #f0f0f0; border-radius: 10px;")
        left_panel_layout = QVBoxLayout(self.memory_left_panel)
        left_panel_layout.setContentsMargins(15, 15, 15, 15)
        left_panel_layout.setSpacing(10)
        left_panel_layout.setAlignment(Qt.AlignTop)

        # Right Panel (Results)
        self.memory_right_panel = QFrame()
        self.memory_right_panel.setStyleSheet(f"background-color: white; border: 1px solid #ccc; border-radius: 10px;")
        right_panel_layout = QVBoxLayout(self.memory_right_panel)
        self.memory_results_view = QTextEdit()
        self.memory_results_view.setReadOnly(True)
        self.memory_results_view.setFont(QFont("Consolas", 10))
        self.memory_results_view.setStyleSheet("border: none; background-color: white; padding: 5px;")
        right_panel_layout.addWidget(self.memory_results_view)

        content_layout.addWidget(self.memory_left_panel)
        content_layout.addWidget(self.memory_right_panel, 1)
        memory_layout.addLayout(content_layout, 1)

        # --- Create Sub-Option Button Groups ---
        self.memory_sub_option_panels = {
            "Core Analysis Files": self._create_sub_option_panel(["virustotal", "filtered netscan with IPcheck", "virustotal IP"]),
            "Volatility": self._create_sub_option_panel(["malfind", "pslist", "netscan", "userassist", "wininfo", "cmdline"]),
            "Memory Dumps": self._create_sub_option_panel(["dumped memory", "dumped memory malicious ips"])
        }
        
        for panel in self.memory_sub_option_panels.values():
            left_panel_layout.addWidget(panel)

        # Set initial state
        self.memory_tabs["Core Analysis Files"].setChecked(True)
        self._on_memory_tab_click()
        
    def _create_sub_option_panel(self, button_names):
        """Creates a container with a set of buttons for the left panel."""
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0,0,0,0)
        layout.setSpacing(10)
        
        buttons = []
        for name in button_names:
            button = QPushButton(name)
            button.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
            button.setCheckable(True)
            button.setStyleSheet(f"""
                QPushButton {{
                    background-color: {COLOR_DARK}; color: white; border: none;
                    border-radius: 8px; padding: 12px; text-align: center;
                }}
                QPushButton:hover {{ background-color: #555; }}
                QPushButton:checked {{ background-color: {COLOR_ORANGE}; }}
            """)
            button.clicked.connect(self._on_memory_sub_option_click)
            layout.addWidget(button)
            buttons.append(button)
        panel.setProperty("buttons", buttons)
        return panel

    def _on_memory_tab_click(self):
        """Handles switching between the main memory analysis tabs."""
        sender = self.sender()
        if not sender: # Initial call
            sender = self.memory_tabs["Core Analysis Files"]

        for name, button in self.memory_tabs.items():
            is_active = (button == sender)
            button.setChecked(is_active)
            self.memory_sub_option_panels[name].setVisible(is_active)
            button.setStyleSheet(f"background-color: {COLOR_ORANGE if is_active else COLOR_DARK}; color: white; border-radius: 8px; padding: 10px;")
        
        self.memory_results_view.clear()
        # Auto-click the first sub-option in the visible panel
        for panel in self.memory_sub_option_panels.values():
            if panel.isVisible():
                buttons = panel.property("buttons")
                if buttons:
                    buttons[0].click()
                break

    def _on_memory_sub_option_click(self):
        """Handles clicks on the sub-option buttons in the left panel."""
        clicked_button = self.sender()
        # Uncheck other buttons in the same panel
        for panel in self.memory_sub_option_panels.values():
            if panel.isVisible():
                for button in panel.property("buttons"):
                    if button != clicked_button:
                        button.setChecked(False)
        
        option_name = clicked_button.text()

        # Default placeholder
        self.memory_results_view.setHtml(f"<h3>Displaying results for: {option_name}</h3><p>(Analysis logic not yet implemented)</p>")

        # Example from image for 'virustotal' with improved formatting
        if option_name == "virustotal":
            html_content = """
            <div style="font-family: Segoe UI, sans-serif; font-size: 10pt; line-height: 1.6;">
                <p>
                    <b>VirusTotal Detections:</b> 45 out of 76 total scans<br/>
                    <b>Malware Status:</b> <span style="color: #d9534f; font-weight: bold;">Malicious</span>
                </p>
                <hr>
                <b style="font-size: 11pt;">Detection Engines and Results:</b>
                <br/><br/>
                <table width="100%" style="border-collapse: collapse; font-size: 9pt;">
                    <tr style="background-color: #f7f7f7;">
                        <td width="30%" style="padding: 8px; border: 1px solid #ddd;"><b>Bkav</b></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">W32.AIDetectMalware.CS</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><b>Lionic</b></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">Trojan.Win32.Generic.4!c</td>
                    </tr>
                    <tr style="background-color: #f7f7f7;">
                        <td style="padding: 8px; border: 1px solid #ddd;"><b>MicroWorld-eScan</b></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">Trojan.GenericKD.74126631</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><b>CTX</b></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">exe.trojan.generic</td>
                    </tr>
                    <tr style="background-color: #f7f7f7;">
                        <td style="padding: 8px; border: 1px solid #ddd;"><b>CAT-QuickHeal</b></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">Trojan.Ghanarava.17306106901e6947</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><b>Skyhigh</b></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">Mars-Stealer!F40168CC10F8</td>
                    </tr>
                    <tr style="background-color: #f7f7f7;">
                        <td style="padding: 8px; border: 1px solid #ddd;"><b>McAfee</b></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">Mars-Stealer!F40168CC10F8</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><b>Cylance</b></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">Unsafe</td>
                    </tr>
                    <tr style="background-color: #f7f7f7;">
                        <td style="padding: 8px; border: 1px solid #ddd;"><b>Sangfor</b></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">Trojan.Win32.Agent.Vnnz9</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><b>Alibaba</b></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">Trojan:Win32/Mindluks.1468e94e</td>
                    </tr>
                    <tr style="background-color: #f7f7f7;">
                        <td style="padding: 8px; border: 1px solid #ddd;"><b>CrowdStrike</b></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">win/malicious_confidence_60% (D)</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><b>Symantec</b></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">Trojan.Gen.MBT</td>
                    </tr>
                    <tr style="background-color: #f7f7f7;">
                        <td style="padding: 8px; border: 1px solid #ddd;"><b>Trellix</b></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">Generic.Malware</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd;"><b>APEX</b></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">Malicious</td>
                    </tr>
                     <tr style="background-color: #f7f7f7;">
                        <td style="padding: 8px; border: 1px solid #ddd;"><b>TrendMicro-HouseCall</b></td>
                        <td style="padding: 8px; border: 1px solid #ddd;">TROJ_GEN.R002C0PII24</td>
                    </tr>
                </table>
            </div>
            """
            self.memory_results_view.setHtml(html_content)

        elif option_name == "malfind":
            # Sample data provided by user
            json_data = [ { "PID": "5896", "Process": "oneetx.exe", "Start VPN": "0x400000", "End VPN": "0x437fff", "Tag": "VadS", "Protection": "PAGE_EXECUTE_READWRITE", "CommitCharge": "56", "PrivateMemory": "1", "File output": "Disabled", "Notes": "MZ header", "Hexdump": [ "4d 5a 90 00 03 00 00 00 04 00 00 00 ff ff 00 00 MZ..............", "b8 00 00 00 00 00 00 00 40 00 00 00 00 00 00 00 ........@.......", "00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 ................", "00 00 00 00 00 00 00 00 00 00 00 00 00 01 00 00 ................" ], "Disasm": [ "0x400000:\tdec\tebp", "0x400001:\tpop\tedx", "0x400002:\tnop\t", "0x400003:\tadd\tbyte ptr [ebx], al", "0x400005:\tadd\tbyte ptr [eax], al", "0x400007:\tadd\tbyte ptr [eax + eax], al", "0x40000a:\tadd\tbyte ptr [eax], al" ] }, { "PID": "7540", "Process": "smartscreen.ex", "Start VPN": "0x2505c140000", "End VPN": "0x2505c15ffff", "Tag": "VadS", "Protection": "PAGE_EXECUTE_READWRITE", "CommitCharge": "1", "PrivateMemory": "1", "File output": "Disabled", "Notes": "N/A", "Hexdump": [ "48 89 54 24 10 48 89 4c 24 08 4c 89 44 24 18 4c H.T$.H.L$.L.D$.L", "89 4c 24 20 48 8b 41 28 48 8b 48 08 48 8b 51 50 .L$ H.A(H.H.H.QP", "48 83 e2 f8 48 8b ca 48 b8 60 00 14 5c 50 02 00 H...H..H.`..\\P..", "00 48 2b c8 48 81 f9 70 0f 00 00 76 09 48 c7 c1 .H+.H..p...v.H.." ], "Disasm": [ "0x2505c140000:\tmov\tqword ptr [rsp + 0x10], rdx", "0x2505c140005:\tmov\tqword ptr [rsp + 8], rcx", "0x2505c14000a:\tmov\tqword ptr [rsp + 0x18], r8", "0x2505c14000f:\tmov\tqword ptr [rsp + 0x20], r9", "0x2505c140014:\tmov\trax, qword ptr [rcx + 0x28]", "0x2505c140018:\tmov\trcx, qword ptr [rax + 8]", "0x2505c14001c:\tmov\trdx, qword ptr [rcx + 0x50]", "0x2505c140020:\tand\trdx, 0xfffffffffffffff8", "0x2505c140024:\tmov\trcx, rdx", "0x2505c140027:\tmovabs\trax, 0x2505c140060", "0x2505c140031:\tsub\trcx, rax", "0x2505c140034:\tcmp\trcx, 0xf70", "0x2505c14003b:\tjbe\t0x2505c140046" ] } ]

            malfind_cards_html = ""
            for region in json_data:
                protection_color = "#d9534f" if "EXECUTE" in region.get("Protection", "") else "#5bc0de"
                hexdump_str = "<br/>".join(region.get("Hexdump", []))
                disasm_str = "<br/>".join(region.get("Disasm", []))

                malfind_cards_html += f"""
                <div style="border: 1px solid #ddd; border-radius: 5px; padding: 15px; margin-bottom: 20px; background-color: #ffffff;">
                    <h4 style="font-size: 11pt; margin-top: 0; background-color: #f0f0f0; padding: 10px; border-radius: 4px;">
                        Process: <b>{region['Process']}</b> (PID: {region['PID']})
                    </h4>
                    <table width="100%" style="font-size: 9pt;">
                        <tr>
                            <td style="padding: 4px; width: 120px;"><b>Start VPN</b></td><td style="padding: 4px;">{region['Start VPN']}</td>
                            <td style="padding: 4px; width: 120px;"><b>End VPN</b></td><td style="padding: 4px;">{region['End VPN']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 4px;"><b>Tag</b></td><td style="padding: 4px;">{region['Tag']}</td>
                            <td style="padding: 4px;"><b>Protection</b></td><td style="padding: 4px;"><span style="color: {protection_color}; font-weight: bold;">{region['Protection']}</span></td>
                        </tr>
                         <tr>
                            <td style="padding: 4px;"><b>CommitCharge</b></td><td style="padding: 4px;">{region['CommitCharge']}</td>
                            <td style="padding: 4px;"><b>Notes</b></td><td style="padding: 4px;">{region['Notes']}</td>
                        </tr>
                    </table>
                    <hr style="margin-top: 15px; margin-bottom: 10px; border: 0; border-top: 1px solid #eee;"/>
                    <table width="100%" style="font-size: 9pt; table-layout: fixed;">
                        <tr style="text-align: left;">
                            <th style="padding: 8px; width: 50%;">Hex Dump</th>
                            <th style="padding: 8px;">Disassembly</th>
                        </tr>
                        <tr>
                            <td style="vertical-align: top; padding: 8px; background-color: #fafafa; border: 1px solid #eee;">
                                <pre style="margin: 0; font-family: Consolas, monaco, monospace; font-size: 8pt; white-space: pre-wrap;">{hexdump_str}</pre>
                            </td>
                            <td style="vertical-align: top; padding: 8px; background-color: #fafafa; border: 1px solid #eee;">
                                <pre style="margin: 0; font-family: Consolas, monaco, monospace; font-size: 8pt; white-space: pre-wrap;">{disasm_str}</pre>
                            </td>
                        </tr>
                    </table>
                </div>
                """
            
            html_content = f"""
            <div style="font-family: Segoe UI, sans-serif; font-size: 10pt; line-height: 1.6; background-color: #f7f7f7; padding: 10px;">
                <h3 style="font-size: 13pt;">Malfind: Find Hidden or Injected Code</h3>
                {malfind_cards_html}
            </div>
            """
            self.memory_results_view.setHtml(html_content)

        elif option_name == "pslist":
            # Sample data provided by user
            json_data = [ { "PID": "4", "PPID": "0", "ImageFileName": "System", "Offset(V)": "0xad8185883180", "Threads": "157", "Handles": "-", "SessionId": "N/A", "Wow64": "False", "CreateTime": "2023-05-21 22:27:10.000000 UTC", "ExitTime": "N/A", "File output": "Disabled" }, { "PID": "108", "PPID": "4", "ImageFileName": "Registry", "Offset(V)": "0xad81858f2080", "Threads": "4", "Handles": "-", "SessionId": "N/A", "Wow64": "False", "CreateTime": "2023-05-21 22:26:54.000000 UTC", "ExitTime": "N/A", "File output": "Disabled" }, { "PID": "332", "PPID": "4", "ImageFileName": "smss.exe", "Offset(V)": "0xad81860dc040", "Threads": "2", "Handles": "-", "SessionId": "N/A", "Wow64": "False", "CreateTime": "2023-05-21 22:27:10.000000 UTC", "ExitTime": "N/A", "File output": "Disabled" }, { "PID": "452", "PPID": "444", "ImageFileName": "csrss.exe", "Offset(V)": "0xad81861cd080", "Threads": "12", "Handles": "-", "SessionId": "0", "Wow64": "False", "CreateTime": "2023-05-21 22:27:22.000000 UTC", "ExitTime": "N/A", "File output": "Disabled" }, { "PID": "528", "PPID": "520", "ImageFileName": "csrss.exe", "Offset(V)": "0xad8186f1b140", "Threads": "14", "Handles": "-", "SessionId": "1", "Wow64": "False", "CreateTime": "2023-05-21 22:27:25.000000 UTC", "ExitTime": "N/A", "File output": "Disabled" }, { "PID": "552", "PPID": "444", "ImageFileName": "wininit.exe", "Offset(V)": "0xad8186f2b080", "Threads": "1", "Handles": "-", "SessionId": "0", "Wow64": "False", "CreateTime": "2023-05-21 22:27:25.000000 UTC", "ExitTime": "N/A", "File output": "Disabled" }, { "PID": "588", "PPID": "520", "ImageFileName": "winlogon.exe", "Offset(V)": "0xad8186f450c0", "Threads": "5", "Handles": "-", "SessionId": "1", "Wow64": "False", "CreateTime": "2023-05-21 22:27:25.000000 UTC", "ExitTime": "N/A", "File output": "Disabled" }, { "PID": "676", "PPID": "552", "ImageFileName": "services.exe", "Offset(V)": "0xad8186f4d080", "Threads": "7", "Handles": "-", "SessionId": "0", "Wow64": "False", "CreateTime": "2023-05-21 22:27:29.000000 UTC", "ExitTime": "N/A", "File output": "Disabled" }, { "PID": "696", "PPID": "552", "ImageFileName": "lsass.exe", "Offset(V)": "0xad8186fc6080", "Threads": "10", "Handles": "-", "SessionId": "0", "Wow64": "False", "CreateTime": "2023-05-21 22:27:29.000000 UTC", "ExitTime": "N/A", "File output": "Disabled" }, { "PID": "824", "PPID": "676", "ImageFileName": "svchost.exe", "Offset(V)": "0xad818761d240", "Threads": "22", "Handles": "-", "SessionId": "0", "Wow64": "False", "CreateTime": "2023-05-21 22:27:32.000000 UTC", "ExitTime": "N/A", "File output": "Disabled" }, { "PID": "852", "PPID": "552", "ImageFileName": "fontdrvhost.ex", "Offset(V)": "0xad818761b0c0", "Threads": "5", "Handles": "-", "SessionId": "0", "Wow64": "False", "CreateTime": "2023-05-21 22:27:33.000000 UTC", "ExitTime": "N/A", "File output": "Disabled" }, { "PID": "860", "PPID": "588", "ImageFileName": "fontdrvhost.ex", "Offset(V)": "0xad818761f140", "Threads": "5", "Handles": "-", "SessionId": "1", "Wow64": "False", "CreateTime": "2023-05-21 22:27:33.000000 UTC", "ExitTime": "N/A", "File output": "Disabled" } ]

            # --- Build HTML for Pslist ---
            table_rows_html = ""
            for i, process in enumerate(json_data):
                row_style = "background-color: #f9f9f9;" if i % 2 == 0 else ""
                table_rows_html += f"""
                <tr style="{row_style}">
                    <td style="padding: 6px; border: 1px solid #eee;">{process['PID']}</td>
                    <td style="padding: 6px; border: 1px solid #eee;">{process['PPID']}</td>
                    <td style="padding: 6px; border: 1px solid #eee; font-family: Consolas, monaco, monospace;">{process['ImageFileName']}</td>
                    <td style="padding: 6px; border: 1px solid #eee; font-family: Consolas, monaco, monospace;">{process['Offset(V)']}</td>
                    <td style="padding: 6px; border: 1px solid #eee;">{process['Threads']}</td>
                    <td style="padding: 6px; border: 1px solid #eee;">{process['Handles']}</td>
                    <td style="padding: 6px; border: 1px solid #eee;">{process['SessionId']}</td>
                    <td style="padding: 6px; border: 1px solid #eee;">{process['Wow64']}</td>
                    <td style="padding: 6px; border: 1px solid #eee;">{process['CreateTime']}</td>
                    <td style="padding: 6px; border: 1px solid #eee;">{process['ExitTime']}</td>
                </tr>
                """
            
            html_content = f"""
            <div style="font-family: Segoe UI, sans-serif; font-size: 10pt; line-height: 1.6;">
                <h3 style="font-size: 13pt;">Pslist: List Running Processes</h3>
                <table width="100%" style="border-collapse: collapse; font-size: 9pt;">
                    <thead>
                        <tr style="background-color: #343a40; color: white; text-align: left;">
                            <th style="padding: 8px; border: 1px solid #ddd;">PID</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">PPID</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Image File Name</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Offset(V)</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Threads</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Handles</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Session ID</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Wow64</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Create Time</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Exit Time</th>
                        </tr>
                    </thead>
                    <tbody>
                        {table_rows_html}
                    </tbody>
                </table>
            </div>
            """
            self.memory_results_view.setHtml(html_content)

        elif option_name == "netscan":
            # Sample data provided by user
            json_data = [ { "Offset": "0xad81861e2310", "Proto": "TCPv4", "LocalAddr": "0.0.0.0", "LocalPort": "49668", "ForeignAddr": "0.0.0.0", "ForeignPort": "0", "State": "LISTENING", "PID": "1840", "Owner": "spoolsv.exe", "Created": "2023-05-21" }, { "Offset": "0xad81861e2310", "Proto": "TCPv6", "LocalAddr": "::", "LocalPort": "49668", "ForeignAddr": "::", "ForeignPort": "0", "State": "LISTENING", "PID": "1840", "Owner": "spoolsv.exe", "Created": "2023-05-21" }, { "Offset": "0xad81861e2470", "Proto": "TCPv4", "LocalAddr": "0.0.0.0", "LocalPort": "5040", "ForeignAddr": "0.0.0.0", "ForeignPort": "0", "State": "LISTENING", "PID": "1196", "Owner": "svchost.exe", "Created": "2023-05-21" }, { "Offset": "0xad81861e2730", "Proto": "TCPv4", "LocalAddr": "0.0.0.0", "LocalPort": "135", "ForeignAddr": "0.0.0.0", "ForeignPort": "0", "State": "LISTENING", "PID": "952", "Owner": "svchost.exe", "Created": "2023-05-21" }, { "Offset": "0xad81861e2b50", "Proto": "TCPv4", "LocalAddr": "0.0.0.0", "LocalPort": "49665", "ForeignAddr": "0.0.0.0", "ForeignPort": "0", "State": "LISTENING", "PID": "552", "Owner": "wininit.exe", "Created": "2023-05-21" }, { "Offset": "0xad81861e2b50", "Proto": "TCPv6", "LocalAddr": "::", "LocalPort": "49665", "ForeignAddr": "::", "ForeignPort": "0", "State": "LISTENING", "PID": "552", "Owner": "wininit.exe", "Created": "2023-05-21" }, { "Offset": "0xad81861e2e10", "Proto": "TCPv4", "LocalAddr": "0.0.0.0", "LocalPort": "49665", "ForeignAddr": "0.0.0.0", "ForeignPort": "0", "State": "LISTENING", "PID": "552", "Owner": "wininit.exe", "Created": "2023-05-21" }, { "Offset": "0xad81861e3230", "Proto": "TCPv4", "LocalAddr": "0.0.0.0", "LocalPort": "49664", "ForeignAddr": "0.0.0.0", "ForeignPort": "0", "State": "LISTENING", "PID": "696", "Owner": "lsass.exe", "Created": "2023-05-21" }, { "Offset": "0xad81861e3390", "Proto": "TCPv4", "LocalAddr": "0.0.0.0", "LocalPort": "135", "ForeignAddr": "0.0.0.0", "ForeignPort": "0", "State": "LISTENING", "PID": "952", "Owner": "svchost.exe", "Created": "2023-05-21" }, { "Offset": "0xad81861e3390", "Proto": "TCPv6", "LocalAddr": "::", "LocalPort": "135", "ForeignAddr": "::", "ForeignPort": "0", "State": "LISTENING", "PID": "952", "Owner": "svchost.exe", "Created": "2023-05-21" }, { "Offset": "0xad81861e34f0", "Proto": "TCPv4", "LocalAddr": "0.0.0.0", "LocalPort": "49664", "ForeignAddr": "0.0.0.0", "ForeignPort": "0", "State": "LISTENING", "PID": "696", "Owner": "lsass.exe", "Created": "2023-05-21" }, { "Offset": "0xad81861e34f0", "Proto": "TCPv6", "LocalAddr": "::", "LocalPort": "49664", "ForeignAddr": "::", "ForeignPort": "0", "State": "LISTENING", "PID": "696", "Owner": "lsass.exe", "Created": "2023-05-21" }, { "Offset": "0xad81861e37b0", "Proto": "TCPv4", "LocalAddr": "0.0.0.0", "LocalPort": "49666", "ForeignAddr": "0.0.0.0", "ForeignPort": "0", "State": "LISTENING", "PID": "1012", "Owner": "svchost.exe", "Created": "2023-05-21" }, { "Offset": "0xad81861e37b0", "Proto": "TCPv6", "LocalAddr": "::", "LocalPort": "49666", "ForeignAddr": "::", "ForeignPort": "0", "State": "LISTENING", "PID": "1012", "Owner": "svchost.exe", "Created": "2023-05-21" }, { "Offset": "0xad81861e3910", "Proto": "TCPv4", "LocalAddr": "0.0.0.0", "LocalPort": "49667", "ForeignAddr": "0.0.0.0", "ForeignPort": "0", "State": "LISTENING", "PID": "448", "Owner": "svchost.exe", "Created": "2023-05-21" } ]

            # --- Build HTML for Netscan ---
            table_rows_html = ""
            for i, connection in enumerate(json_data):
                row_style = "background-color: #f9f9f9;" if i % 2 == 0 else ""
                local_addr = f"{connection['LocalAddr']}:{connection['LocalPort']}"
                foreign_addr = f"{connection['ForeignAddr']}:{connection['ForeignPort']}"
                
                # Color code the state for better visibility
                state_color = "#28a745" if connection['State'] == "LISTENING" else "#ffc107"
                
                table_rows_html += f"""
                <tr style="{row_style}">
                    <td style="padding: 6px; border: 1px solid #eee; font-family: Consolas, monaco, monospace;">{connection['Offset']}</td>
                    <td style="padding: 6px; border: 1px solid #eee;">{connection['Proto']}</td>
                    <td style="padding: 6px; border: 1px solid #eee; font-family: Consolas, monaco, monospace;">{local_addr}</td>
                    <td style="padding: 6px; border: 1px solid #eee; font-family: Consolas, monaco, monospace;">{foreign_addr}</td>
                    <td style="padding: 6px; border: 1px solid #eee; color: {state_color}; font-weight: bold;">{connection['State']}</td>
                    <td style="padding: 6px; border: 1px solid #eee;">{connection['PID']}</td>
                    <td style="padding: 6px; border: 1px solid #eee; font-family: Consolas, monaco, monospace;">{connection['Owner']}</td>
                    <td style="padding: 6px; border: 1px solid #eee;">{connection['Created']}</td>
                </tr>
                """
            
            html_content = f"""
            <div style="font-family: Segoe UI, sans-serif; font-size: 10pt; line-height: 1.6;">
                <h3 style="font-size: 13pt;">Netscan: Network Connections</h3>
                <table width="100%" style="border-collapse: collapse; font-size: 9pt;">
                    <thead>
                        <tr style="background-color: #343a40; color: white; text-align: left;">
                            <th style="padding: 8px; border: 1px solid #ddd;">Offset</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Protocol</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Local Address</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Foreign Address</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">State</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">PID</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Owner</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Created</th>
                        </tr>
                    </thead>
                    <tbody>
                        {table_rows_html}
                    </tbody>
                </table>
            </div>
            """
            self.memory_results_view.setHtml(html_content)

        elif option_name == "userassist":
            # Sample data provided by user
            json_data = [ { "PID": "4", "Process": "System", "UserAssist": "N/A", "LastUsed": "N/A", "LastUsedTime": "N/A" }, { "PID": "108", "Process": "Registry", "UserAssist": "N/A", "LastUsed": "N/A", "LastUsedTime": "N/A" }, { "PID": "332", "Process": "smss.exe", "UserAssist": "N/A", "LastUsed": "N/A", "LastUsedTime": "N/A" }, { "PID": "452", "Process": "csrss.exe", "UserAssist": "N/A", "LastUsed": "N/A", "LastUsedTime": "N/A" }, { "PID": "528", "Process": "csrss.exe", "UserAssist": "N/A", "LastUsed": "N/A", "LastUsedTime": "N/A" }, { "PID": "552", "Process": "wininit.exe", "UserAssist": "N/A", "LastUsed": "N/A", "LastUsedTime": "N/A" }, { "PID": "588", "Process": "winlogon.exe", "UserAssist": "N/A", "LastUsed": "N/A", "LastUsedTime": "N/A" }, { "PID": "676", "Process": "services.exe", "UserAssist": "N/A", "LastUsed": "N/A", "LastUsedTime": "N/A" }, { "PID": "696", "Process": "lsass.exe", "UserAssist": "N/A", "LastUsed": "N/A", "LastUsedTime": "N/A" }, { "PID": "824", "Process": "svchost.exe", "UserAssist": "N/A", "LastUsed": "N/A", "LastUsedTime": "N/A" }, { "PID": "852", "Process": "fontdrvhost.ex", "UserAssist": "N/A", "LastUsed": "N/A", "LastUsedTime": "N/A" }, { "PID": "860", "Process": "fontdrvhost.ex", "UserAssist": "N/A", "LastUsed": "N/A", "LastUsedTime": "N/A" } ]

            # --- Build HTML for UserAssist ---
            table_rows_html = ""
            for i, process in enumerate(json_data):
                row_style = "background-color: #f9f9f9;" if i % 2 == 0 else ""
                table_rows_html += f"""
                <tr style="{row_style}">
                    <td style="padding: 6px; border: 1px solid #eee;">{process['PID']}</td>
                    <td style="padding: 6px; border: 1px solid #eee; font-family: Consolas, monaco, monospace;">{process['Process']}</td>
                    <td style="padding: 6px; border: 1px solid #eee;">{process['UserAssist']}</td>
                    <td style="padding: 6px; border: 1px solid #eee;">{process['LastUsed']}</td>
                    <td style="padding: 6px; border: 1px solid #eee;">{process['LastUsedTime']}</td>
                </tr>
                """
            
            html_content = f"""
            <div style="font-family: Segoe UI, sans-serif; font-size: 10pt; line-height: 1.6;">
                <h3 style="font-size: 13pt;">UserAssist: Last Used Programs</h3>
                <table width="100%" style="border-collapse: collapse; font-size: 9pt;">
                    <thead>
                        <tr style="background-color: #343a40; color: white; text-align: left;">
                            <th style="padding: 8px; border: 1px solid #ddd;">PID</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Process</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">UserAssist</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Last Used</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Last Used Time</th>
                        </tr>
                    </thead>
                    <tbody>
                        {table_rows_html}
                    </tbody>
                </table>
            </div>
            """
            self.memory_results_view.setHtml(html_content)

        elif option_name == "wininfo":
            # Sample data provided by user
            json_data = [ { "Variable": "Kernel", "Value": "Base" }, { "Variable": "DTB", "Value": "0x1ad000" }, { "Variable": "Symbols", "Value": "file:///D:/Forensics%20tools/volatility3/volatility3/symbols/windows/ntkrnlmp.pdb/68A17FAF3012B7846079AEECDBE0A583-1.json.xz" }, { "Variable": "Is64Bit", "Value": "True" }, { "Variable": "IsPAE", "Value": "False" }, { "Variable": "layer_name", "Value": "0" }, { "Variable": "memory_layer", "Value": "1" }, { "Variable": "KdVersionBlock", "Value": "0xf80762e29398" }, { "Variable": "Major/Minor", "Value": "15.19041" }, { "Variable": "MachineType", "Value": "34404" }, { "Variable": "KeNumberProcessors", "Value": "4" }, { "Variable": "SystemTime", "Value": "2023-05-21" }, { "Variable": "NtSystemRoot", "Value": "C:\\Windows" }, { "Variable": "NtProductType", "Value": "NtProductWinNt" }, { "Variable": "NtMajorVersion", "Value": "10" }, { "Variable": "NtMinorVersion", "Value": "0" }, { "Variable": "PE", "Value": "MajorOperatingSystemVersion" }, { "Variable": "PE", "Value": "MinorOperatingSystemVersion" }, { "Variable": "PE", "Value": "Machine" }, { "Variable": "PE", "Value": "TimeDateStamp" } ]

            # --- Build HTML for Wininfo ---
            table_rows_html = ""
            for i, info in enumerate(json_data):
                row_style = "background-color: #f9f9f9;" if i % 2 == 0 else ""
                
                # Format the value based on the variable type
                value = info['Value']
                if info['Variable'] in ['Is64Bit', 'IsPAE']:
                    # Boolean values - color code them
                    value_color = "#28a745" if value == "True" else "#dc3545"
                    value = f'<span style="color: {value_color}; font-weight: bold;">{value}</span>'
                elif info['Variable'] in ['DTB', 'KdVersionBlock']:
                    # Hexadecimal values - use monospace font
                    value = f'<span style="font-family: Consolas, monaco, monospace;">{value}</span>'
                elif info['Variable'] == 'Symbols':
                    # Long path - truncate and show in monospace
                    if len(value) > 80:
                        value = f'<span style="font-family: Consolas, monaco, monospace;" title="{value}">{value[:80]}...</span>'
                    else:
                        value = f'<span style="font-family: Consolas, monaco, monospace;">{value}</span>'
                elif info['Variable'] == 'NtSystemRoot':
                    # System path - use monospace font
                    value = f'<span style="font-family: Consolas, monaco, monospace;">{value}</span>'
                
                table_rows_html += f"""
                <tr style="{row_style}">
                    <td style="padding: 8px; border: 1px solid #eee; font-weight: bold; width: 200px;">{info['Variable']}</td>
                    <td style="padding: 8px; border: 1px solid #eee;">{value}</td>
                </tr>
                """
            
            html_content = f"""
            <div style="font-family: Segoe UI, sans-serif; font-size: 10pt; line-height: 1.6;">
                <h3 style="font-size: 13pt;">Wininfo: Windows System Information</h3>
                <table width="100%" style="border-collapse: collapse; font-size: 9pt;">
                    <thead>
                        <tr style="background-color: #343a40; color: white; text-align: left;">
                            <th style="padding: 10px; border: 1px solid #ddd;">System Variable</th>
                            <th style="padding: 10px; border: 1px solid #ddd;">Value</th>
                        </tr>
                    </thead>
                    <tbody>
                        {table_rows_html}
                    </tbody>
                </table>
            </div>
            """
            self.memory_results_view.setHtml(html_content)

        elif option_name == "cmdline":
            # Sample data provided by user
            json_data = [ { "PID": "4", "Process": "System", "CommandLine": "N/A" }, { "PID": "108", "Process": "Registry", "CommandLine": "N/A" }, { "PID": "332", "Process": "smss.exe", "CommandLine": "N/A" }, { "PID": "452", "Process": "csrss.exe", "CommandLine": "N/A" }, { "PID": "528", "Process": "csrss.exe", "CommandLine": "N/A" }, { "PID": "552", "Process": "wininit.exe", "CommandLine": "N/A" }, { "PID": "588", "Process": "winlogon.exe", "CommandLine": "N/A" }, { "PID": "676", "Process": "services.exe", "CommandLine": "N/A" }, { "PID": "696", "Process": "lsass.exe", "CommandLine": "N/A" }, { "PID": "824", "Process": "svchost.exe", "CommandLine": "N/A" }, { "PID": "852", "Process": "fontdrvhost.ex", "CommandLine": "N/A" }, { "PID": "860", "Process": "fontdrvhost.ex", "CommandLine": "N/A" } ]

            # --- Build HTML for Command Line ---
            table_rows_html = ""
            for i, process in enumerate(json_data):
                row_style = "background-color: #f9f9f9;" if i % 2 == 0 else ""
                table_rows_html += f"""
                <tr style="{row_style}">
                    <td style="padding: 6px; border: 1px solid #eee;">{process['PID']}</td>
                    <td style="padding: 6px; border: 1px solid #eee; font-family: Consolas, monaco, monospace;">{process['Process']}</td>
                    <td style="padding: 6px; border: 1px solid #eee;">{process['CommandLine']}</td>
                </tr>
                """
            
            html_content = f"""
            <div style="font-family: Segoe UI, sans-serif; font-size: 10pt; line-height: 1.6;">
                <h3 style="font-size: 13pt;">Command Line: Last Used Programs</h3>
                <table width="100%" style="border-collapse: collapse; font-size: 9pt;">
                    <thead>
                        <tr style="background-color: #343a40; color: white; text-align: left;">
                            <th style="padding: 8px; border: 1px solid #ddd;">PID</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Process</th>
                            <th style="padding: 8px; border: 1px solid #ddd;">Command Line</th>
                        </tr>
                    </thead>
                    <tbody>
                        {table_rows_html}
                    </tbody>
                </table>
            </div>
            """
            self.memory_results_view.setHtml(html_content)

    def on_artifact_button_click(self, artifact_name):
<<<<<<< HEAD
        """Handle clicks on the left-side artifact buttons."""
=======
        """Handle clicks on the artifact buttons."""
        self._switch_right_panel_view(self.placeholder_label)
        self.placeholder_label.setText(f"Gathering data for {artifact_name}...")

        if artifact_name == "MEMORY":
            self._switch_right_panel_view(self.memory_view_container)
            return

>>>>>>> 99e81c0b01190573585baf4d5569a60786b8c6be
        if artifact_name == "WEB":
            if not self.connection_params:
                QMessageBox.warning(self, "No Connection", "Please establish a remote connection first.")
                return
            self.web_view.load(QUrl()) # Clear previous content
            self._switch_right_panel_view(self.web_view)
            self.web_artifact_thread = WebArtifactThread(self.connection_params)
            self.web_artifact_thread.finished.connect(self.on_web_extraction_finished)
            self.web_artifact_thread.start()
        elif artifact_name == "USB":
            self._switch_right_panel_view(self.usb_view_container)
            self.usb_device_thread = UsbDeviceThread()
            self.usb_device_thread.finished.connect(self.on_usb_scan_finished)
            self.usb_device_thread.start()
<<<<<<< HEAD
        elif artifact_name == "REGISTRY":
            if not self.selected_case_path:
                QMessageBox.warning(self, "No Case Selected", "A case must be selected to perform registry analysis.")
                return
            # Update output paths before showing
            self.set_case_path(self.selected_case_path)
            self._switch_right_panel_view(self.registry_view_container)
=======

        elif artifact_name == "SRUM":
            self.start_srum_analysis()
        
>>>>>>> 99e81c0b01190573585baf4d5569a60786b8c6be
        else:
            self.placeholder_label.setText(f"{artifact_name} analysis not implemented yet.")
            self._switch_right_panel_view(self.placeholder_label)

    def start_srum_analysis(self):
        """Initiates the SRUM analysis process with hardcoded paths."""
        if not SRUM_IMPORTS_AVAILABLE:
            QMessageBox.critical(self, "Missing Dependencies", "Required libraries for SRUM analysis (pyesedb, openpyxl, python-registry) are not installed.")
            self.placeholder_label.setText("SRUM analysis requires additional libraries.")
            return

        # Hardcoded paths as requested
        script_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.join(script_dir, "..")
        srum_path = os.path.join(project_root, "SRUDB.dat")
        reg_path = os.path.join(project_root, "SOFTWARE")
        template_path = os.path.join(project_root, "SRUM_TEMPLATE2.XLSX")  # In project root directory

        # --- File Existence Checks ---
        if not os.path.exists(srum_path):
            QMessageBox.critical(self, "File Not Found", f"The required SRUDB.dat file was not found at:\n{srum_path}")
            self.placeholder_label.setText("SRUDB.dat not found at the specified path.")
            return

        if not os.path.exists(template_path):
            QMessageBox.critical(self, "File Not Found", f"The required SRUM template file was not found at:\n{template_path}")
            self.placeholder_label.setText(f"SRUM template not found at {template_path}.")
            return

        # The registry hive is optional, but we check for it since a path is provided.
        if not os.path.exists(reg_path):
            QMessageBox.warning(self, "Registry Hive Not Found", 
                                f"The SOFTWARE hive was not found at the specified path:\n{reg_path}\n\nAnalysis will proceed without it.")
            reg_path = None  # Set to None so the analyzer can handle it gracefully
        
        params = {
            "srum_path": srum_path,
            "template_path": template_path,
            "reg_path": reg_path
        }

        self.placeholder_label.setText("Analyzing SRUM database... This may take a while.")
        self.srum_analysis_thread = SrumAnalysisThread(params)
        self.srum_analysis_thread.finished.connect(self.on_srum_analysis_finished)
        self.srum_analysis_thread.start()

    def on_srum_analysis_finished(self, result):
        """Handles the finished signal from the SRUM analysis thread."""
        if result["status"] == "success":
            self.display_srum_data(result['data'])
            self._switch_right_panel_view(self.srum_tab_widget)
        else:
            self.placeholder_label.setText(f"SRUM Analysis Error: {result['message']}")
            self._switch_right_panel_view(self.placeholder_label)
            QMessageBox.critical(self, "SRUM Analysis Failed", result['message'])

    def display_srum_data(self, all_tables_data):
        """Populates the tab widget with SRUM data with enhanced organization and styling."""
        self.srum_tab_widget.clear()
        if not all_tables_data:
            self.placeholder_label.setText("No data found in SRUM database.")
            self._switch_right_panel_view(self.placeholder_label)
            return

        for tname, table_data in all_tables_data.items():
            if not table_data or len(table_data) < 2:  # Skip empty or header-only tables
                continue

            tab = QWidget()
            layout = QVBoxLayout(tab)
            layout.setContentsMargins(10, 10, 10, 10)
            layout.setSpacing(10)

            # Add header with table info
            header_frame = QFrame()
            header_frame.setStyleSheet("background-color: #f8f9fa; border: 1px solid #dee2e6; border-radius: 5px;")
            header_layout = QHBoxLayout(header_frame)
            
            # Table info
            info_label = QLabel(f"<b>Table:</b> {tname} | <b>Records:</b> {len(table_data) - 1}")
            info_label.setFont(QFont("Segoe UI", 10))
            header_layout.addWidget(info_label)
            
            # Export button
            export_btn = QPushButton("Export to CSV")
            export_btn.setStyleSheet("""
                QPushButton {
                    background-color: #28a745; color: white; border: none;
                    border-radius: 4px; padding: 6px 12px; font-weight: bold;
                }
                QPushButton:hover { background-color: #218838; }
            """)
            export_btn.clicked.connect(lambda checked, data=table_data, name=tname: self.export_srum_csv(data, name))
            header_layout.addWidget(export_btn)
            
            layout.addWidget(header_frame)

            # Add search functionality
            search_frame = QFrame()
            search_frame.setStyleSheet("background-color: #ffffff; border: 1px solid #dee2e6; border-radius: 5px;")
            search_layout = QHBoxLayout(search_frame)
            
            search_label = QLabel("Search:")
            search_label.setFont(QFont("Segoe UI", 9))
            search_layout.addWidget(search_label)
            
            search_box = QLineEdit()
            search_box.setPlaceholderText("Type to filter table data...")
            search_box.setStyleSheet("""
                QLineEdit {
                    border: 1px solid #ced4da; border-radius: 4px; padding: 6px;
                    font-family: 'Segoe UI'; font-size: 9pt;
                }
                QLineEdit:focus { border-color: #80bdff; }
            """)
            search_layout.addWidget(search_box, 1)
            
            layout.addWidget(search_frame)

            # Create enhanced table
            table = QTableWidget()
            table.setSortingEnabled(True)
            table.setAlternatingRowColors(True)
            table.setSelectionBehavior(QTableWidget.SelectRows)
            table.setEditTriggers(QTableWidget.NoEditTriggers)
            table.verticalHeader().setVisible(False)
            table.setStyleSheet("""
                QTableWidget {
                    gridline-color: #dee2e6; background-color: white;
                    alternate-background-color: #f8f9fa;
                    font-family: 'Segoe UI'; font-size: 9pt;
                }
                QTableWidget::item {
                    padding: 6px; border-bottom: 1px solid #dee2e6;
                }
                QTableWidget::item:selected {
                    background-color: #007bff; color: white;
                }
                QHeaderView::section {
                    background-color: #343a40; color: white; padding: 8px;
                    border: none; font-weight: bold; font-family: 'Segoe UI';
                }
                QHeaderView::section:hover {
                    background-color: #495057;
                }
            """)

            headings = table_data[0]
            table.setColumnCount(len(headings))
            table.setHorizontalHeaderLabels(headings)
            table.setRowCount(len(table_data) - 1)

            # Populate table with data
            for row_idx, row_data in enumerate(table_data[1:]):
                for col_idx, cell_data in enumerate(row_data):
                    item = QTableWidgetItem(str(cell_data))
                    
                    # Apply special formatting based on content
                    if cell_data and isinstance(cell_data, str):
                        # Format timestamps
                        if any(time_indicator in cell_data.lower() for time_indicator in ['utc', 'gmt', '2023', '2024']):
                            item.setBackground(QColor(255, 248, 220))  # Light yellow for timestamps
                        # Format hex values
                        elif cell_data.startswith('0x') or (len(cell_data) == 8 and all(c in '0123456789abcdefABCDEF' for c in cell_data)):
                            item.setFont(QFont("Consolas", 9))
                            item.setBackground(QColor(240, 248, 255))  # Light blue for hex
                        # Format file paths
                        elif '\\' in cell_data or '/' in cell_data:
                            item.setFont(QFont("Consolas", 9))
                            item.setBackground(QColor(245, 245, 245))  # Light gray for paths
                    
                    table.setItem(row_idx, col_idx, item)
            
            # Auto-resize columns with smart sizing
            table.resizeColumnsToContents()
            table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            
            # Set minimum column widths
            for col in range(table.columnCount()):
                if table.columnWidth(col) < 100:
                    table.setColumnWidth(col, 100)
                elif table.columnWidth(col) > 300:
                    table.setColumnWidth(col, 300)
            
            # Connect search functionality
            search_box.textChanged.connect(lambda text, t=table: self.filter_srum_table(t, text))
            
            layout.addWidget(table, 1)  # Give table most of the space
            
            # Add status bar
            status_bar = QStatusBar()
            status_bar.setStyleSheet("background-color: #f8f9fa; border-top: 1px solid #dee2e6;")
            status_label = QLabel(f"Showing {len(table_data) - 1} records")
            status_label.setFont(QFont("Segoe UI", 9))
            status_bar.addWidget(status_label)
            layout.addWidget(status_bar)
            
            self.srum_tab_widget.addTab(tab, tname)

    def filter_srum_table(self, table, search_text):
        """Filters the SRUM table based on search text."""
        search_text = search_text.lower()
        for row in range(table.rowCount()):
            row_visible = False
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item and search_text in item.text().lower():
                    row_visible = True
                    break
            table.setRowHidden(row, not row_visible)
        
        # Update status
        visible_count = sum(1 for row in range(table.rowCount()) if not table.isRowHidden(row))
        status_bar = table.parent().findChild(QStatusBar)
        if status_bar:
            status_label = status_bar.findChild(QLabel)
            if status_label:
                status_label.setText(f"Showing {visible_count} of {table.rowCount()} records")

    def export_srum_csv(self, table_data, table_name):
        """Exports SRUM table data to CSV file."""
        if not table_data or len(table_data) < 2:
            QMessageBox.warning(self, "Export Failed", "No data to export.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            f"Export {table_name} to CSV", 
            f"SRUM_{table_name.replace(' ', '_')}.csv", 
            "CSV Files (*.csv);;All Files (*)"
        )
        
        if file_path:
            try:
                with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    # Write all rows including header
                    for row in table_data:
                        writer.writerow(row)
                QMessageBox.information(self, "Export Successful", f"SRUM data exported to {file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Export Failed", f"Failed to export to CSV: {e}")

    def on_web_extraction_finished(self, result):
        """Handle finished signal from the web artifact extraction thread."""
        if result["status"] == "success":
            self._switch_right_panel_view(self.web_view)
            self.web_view.setUrl(QUrl.fromLocalFile(result["report_path"]))
        else:
            self._switch_right_panel_view(self.placeholder_label)
            self.placeholder_label.setText(f"Error: {result['message']}")
            QMessageBox.critical(self, "Extraction Failed", result['message'])

    def on_usb_scan_finished(self, devices):
        """Handles the finished signal from the USB device scan thread."""
        self.usb_progress_bar.setVisible(False)
        self.usb_status_bar.clearMessage()
        self.usb_devices = devices
        if not devices:
            self.placeholder_label.setText("No USB devices found or failed to read registry.")
            self._switch_right_panel_view(self.placeholder_label)
            return
        
        self.apply_usb_filters()
        self._switch_right_panel_view(self.usb_view_container)

    def apply_usb_filters(self):
        """Filters and displays USB devices based on search and time criteria."""
        if not self.usb_devices:
            return

        search_term = self.usb_search_box.text().lower()
        time_filter = self.usb_time_filter.currentText()
        
        now = datetime.utcnow()
        time_delta = None
        if time_filter == "Last 7 Days": time_delta = timedelta(days=7)
        elif time_filter == "Last 30 Days": time_delta = timedelta(days=30)
        elif time_filter == "Last 90 Days": time_delta = timedelta(days=90)
        elif time_filter == "Last Year": time_delta = timedelta(days=365)
        
        cutoff_time = now - time_delta if time_delta else None

        filtered_devices = []
        for device in self.usb_devices:
            # Time filter
            if cutoff_time and (not device.get("datetime_obj") or device["datetime_obj"] < cutoff_time):
                continue

            # Search filter
            if search_term:
                matches = False
                for value in device.values():
                    if search_term in str(value).lower():
                        matches = True
                        break
                if not matches:
                    continue
            
            filtered_devices.append(device)

        self.display_usb_data(filtered_devices)

    def display_usb_data(self, devices):
        """Populates the table with a list of USB devices."""
        self.displayed_usb_devices = devices # Store for the details view
        self.usb_status_bar.showMessage("Populating view...")
        self.usb_table_view.setSortingEnabled(False)
        self.usb_table_view.clear()
        
        headers = ["Forensic ID", "Description", "Hardware ID", "Plug-in Time", "Duration", "Manufacturer"]
        self.usb_table_view.setColumnCount(len(headers))
        self.usb_table_view.setHorizontalHeaderLabels(headers)

        self.usb_table_view.setRowCount(len(devices))
        for row, device in enumerate(devices):
            # Populate with new, richer data
            col_data = [
                device.get("Forensic ID", ""), device.get("Description", ""), device.get("Hardware ID", ""),
                device.get("Plug-in Time", ""), device.get("Duration", ""), device.get("Manufacturer", "")
            ]
            for col, value in enumerate(col_data):
                item = QTableWidgetItem(str(value))
                item.setFont(QFont("Segoe UI", 9))
                self.usb_table_view.setItem(row, col, item)
                
        self.usb_table_view.resizeColumnsToContents()
        self.usb_table_view.setSortingEnabled(True)
        connected_count = sum(1 for d in devices if d.get("Connected") == "Yes")
        self.usb_device_count_label.setText(f"{len(devices)} devices found ({connected_count} connected)")
        self.usb_status_bar.clearMessage()
    
    def show_usb_device_details(self, index):
        """Displays a dialog with detailed forensic info for the selected USB device."""
        if index.row() >= len(self.displayed_usb_devices):
            return

        device = self.displayed_usb_devices[index.row()]
        
        details_html = "<h3>Forensic Details</h3><ul>"
        for key, value in sorted(device.items()):
            if key != "datetime_obj": # Don't show the internal datetime object
                details_html += f"<li><b>{key.replace('_', ' ').title()}:</b> {value}</li>"
        details_html += "</ul>"

        QMessageBox.information(self, f"Details for {device.get('Description', 'Device')}", details_html)

    def export_usb_csv(self):
        """Exports the current USB device list to a CSV file."""
        if not self.displayed_usb_devices:
            QMessageBox.warning(self, "Export Failed", "No USB devices to export.")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "Export USB Devices", "", "CSV Files (*.csv);;All Files (*)")
        if file_path:
            try:
                with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(["Forensic ID", "Description", "Hardware ID", "Plug-in Time", "Duration", "Manufacturer"])
                    for device in self.displayed_usb_devices:
                        writer.writerow([
                            device.get("Forensic ID", ""),
                            device.get("Description", ""),
                            device.get("Hardware ID", ""),
                            device.get("Plug-in Time", ""),
                            device.get("Duration", ""),
                            device.get("Manufacturer", "")
                        ])
                QMessageBox.information(self, "Export Successful", f"USB devices exported to {file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Export Failed", f"Failed to export to CSV: {e}")

    def identify_suspicious_patterns(self, devices):
        """Placeholder for identifying suspicious patterns in USB device data."""
        # This method is called by the new forensic button.
        # It should contain the logic to analyze the devices and identify patterns.
        # For now, it just returns a placeholder message.
        QMessageBox.information(self, "Suspicious Patterns", "Suspicious patterns analysis not yet implemented.")

    def perform_forensic_analysis(self):
        """Placeholder for generating and showing/exporting forensic report."""
        # This method is called by the new forensic button.
        # It should contain the logic to generate a comprehensive forensic report.
        # For now, it just returns a placeholder message.
        QMessageBox.information(self, "Forensic Analysis", "Forensic analysis not yet implemented.")

    def show_usb_context_menu(self, position):
        """Shows a context menu for the selected USB device."""
        index = self.usb_table_view.indexAt(position)
        if not index.isValid():
            return

        device = self.displayed_usb_devices[index.row()]
        menu = QMenu()
        menu.setFont(QFont("Segoe UI", 9))

        copy_cell_action = QAction("Copy Cell", self)
        copy_cell_action.setFont(QFont("Segoe UI", 9))
        copy_cell_action.triggered.connect(lambda: self.copy_cell_to_clipboard(index.row(), index.column()))
        menu.addAction(copy_cell_action)

        copy_row_action = QAction("Copy Row", self)
        copy_row_action.setFont(QFont("Segoe UI", 9))
        copy_row_action.triggered.connect(lambda: self.copy_row_to_clipboard(index.row()))
        menu.addAction(copy_row_action)

        menu.exec_(self.usb_table_view.viewport().mapToGlobal(position))

    def copy_cell_to_clipboard(self, row, column):
        """Copies the content of a specific cell to the clipboard."""
        item = self.usb_table_view.item(row, column)
        if item:
            QApplication.clipboard().setText(item.text())

    def copy_row_to_clipboard(self, row):
        """Copies the content of a specific row to the clipboard."""
        # This method is called by the context menu.
        # It should copy the entire row's data.
        # For now, it just copies the first column (Forensic ID) as a placeholder.
        # A more robust solution would involve copying all columns.
        row_data = []
        for col in range(self.usb_table_view.columnCount()):
            item = self.usb_table_view.item(row, col)
            if item:
                row_data.append(item.text())
        QApplication.clipboard().setText("\n".join(row_data))

    def _handle_tab_click(self, clicked_button):
        """Handle tab button clicks"""
        tab_text = clicked_button.text()
        # Potentially navigate to other pages. For now, just print.
        print(f"Tab clicked: {tab_text}")
        super()._handle_tab_click(clicked_button) 

    # --- REGISTRY ANALYSIS METHODS ---
    def create_registry_view(self):
        """Creates the entire view for Registry Analysis options."""
        container = QFrame()
        container.setStyleSheet("background-color: transparent;")
        
        # Main content area
        content_layout = QHBoxLayout(container)
        
        # Left panel - Options
        left_panel = self.create_registry_options_panel()
        content_layout.addWidget(left_panel, 1)
        
        # Right panel - Progress and Results
        right_panel = self.create_registry_progress_panel()
        content_layout.addWidget(right_panel, 1)
        
        return container

    def create_registry_options_panel(self):
        """Create the left panel with registry analysis options"""
        panel = QScrollArea()
        panel.setWidgetResizable(True)
        panel.setStyleSheet("background: white; border-radius: 12px; padding: 10px; border: none;")
        
        content_widget = QWidget()
        panel.setWidget(content_widget)
        
        layout = QVBoxLayout(content_widget)
        
        # Title
        title = QLabel("Registry Analysis Options")
        title.setFont(QFont("Cascadia Mono", 16, QFont.Weight.Bold))
        title.setStyleSheet(f"color: {COLOR_DARK}; margin-bottom: 15px;")
        layout.addWidget(title)
        
        # Option 1: Acquire Registry Hives
        acquire_group = self.create_acquire_hives_group()
        layout.addWidget(acquire_group)
        
        # Option 2: Analyze Registry Hives
        analyze_group = self.create_analyze_hives_group()
        layout.addWidget(analyze_group)
        
        # Option 3: Compare Registry Hives
        compare_group = self.create_compare_hives_group()
        layout.addWidget(compare_group)
        
        # Option 4: Apply Transaction Logs
        logs_group = self.create_apply_logs_group()
        layout.addWidget(logs_group)
        
        # Option 5: Parse Hive Header
        header_group = self.create_parse_header_group()
        layout.addWidget(header_group)
        
        layout.addStretch()
        return panel

    def _get_group_box_style(self):
        return f"""
            QGroupBox {{
                border: 2px solid {COLOR_DARK};
                border-radius: 8px;
                margin-top: 15px;
                padding: 15px;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 10px;
                padding: 0 5px 0 5px;
                color: {COLOR_DARK};
                font-size: 14px;
                font-weight: bold;
            }}
        """

    def _create_small_browse_button(self, callback):
        browse_btn = QPushButton("Browse...")
        browse_btn.setFixedSize(100, 44)
        # Create a smaller version of the standard button style
        small_button_style = self.get_button_style(bg_color=COLOR_DARK, text_color="white", hover_color=COLOR_ORANGE)
        small_button_style = small_button_style.replace("padding: 18px 64px;", "padding: 8px 12px;")
        small_button_style = small_button_style.replace("font-size: 22px;", "font-size: 14px;")
        browse_btn.setStyleSheet(small_button_style)
        browse_btn.clicked.connect(callback)
        return browse_btn
        
    def create_registry_progress_panel(self):
        panel = QWidget()
        panel.setStyleSheet("background: white; border-radius: 12px; padding: 20px;")
        layout = QVBoxLayout(panel)
        title = QLabel("Progress & Results")
        title.setFont(QFont("Cascadia Mono", 18, QFont.Weight.Bold))
        title.setStyleSheet(f"color: {COLOR_DARK}; margin-bottom: 20px;")
        layout.addWidget(title)
        
        self.registry_progress_text = QTextEdit()
        self.registry_progress_text.setReadOnly(True)
        self.registry_progress_text.setStyleSheet(f"""
            QTextEdit {{
                border: 2px solid {COLOR_DARK};
                border-radius: 8px;
                padding: 10px;
                font-family: 'Cascadia Mono';
                font-size: 12px;
                background-color: #f8f8f8;
            }}
        """)
        layout.addWidget(self.registry_progress_text)
        return panel

    def create_acquire_hives_group(self):
        group = QGroupBox("1. Acquire Registry Hives")
        group.setFont(QFont("Cascadia Mono", 12, QFont.Weight.Bold))
        group.setStyleSheet(self._get_group_box_style())
        
        layout = QVBoxLayout(group)
        layout.setSpacing(10)
        
        layout.addWidget(QLabel("Username (optional):"))
        self.username_input = self.create_styled_input("For NTUSER.DAT and UsrClass.dat")
        layout.addWidget(self.username_input)
        
        layout.addSpacing(10)
        layout.addWidget(QLabel("Select Hives to Acquire:"))
        self.hive_list = QListWidget()
        self.hive_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        self.hive_list.setMaximumHeight(150)
        self.hive_list.setStyleSheet(f"border: 1px solid {COLOR_DARK}; border-radius: 5px; padding: 5px;")
        
        available_hives = self.registry_analyzer.get_available_hives()
        for hive in available_hives:
            self.hive_list.addItem(QListWidgetItem(hive))
        layout.addWidget(self.hive_list)
        
        layout.addSpacing(10)
        layout.addWidget(QLabel("Output Directory:"))
        output_layout = QHBoxLayout()
        self.acquire_output_dir_input = self.create_styled_input()
        browse_btn = self._create_small_browse_button(lambda: self.browse_directory(self.acquire_output_dir_input))
        output_layout.addWidget(self.acquire_output_dir_input)
        output_layout.addWidget(browse_btn)
        layout.addLayout(output_layout)
        
        layout.addSpacing(15)
        acquire_btn = self.create_styled_button("Acquire Hives", self.acquire_hives)
        layout.addWidget(acquire_btn, alignment=Qt.AlignCenter)
        
        return group

    def create_analyze_hives_group(self):
        group = QGroupBox("2. Analyze Registry Hives")
        group.setFont(QFont("Cascadia Mono", 12, QFont.Weight.Bold))
        group.setStyleSheet(self._get_group_box_style())
        layout = QVBoxLayout(group)
        layout.setSpacing(10)

        layout.addWidget(QLabel("Directory of Acquired Hives:"))
        input_layout = QHBoxLayout()
        self.analyze_input_dir = self.create_styled_input()
        browse_input_btn = self._create_small_browse_button(lambda: self.browse_directory(self.analyze_input_dir))
        input_layout.addWidget(self.analyze_input_dir)
        input_layout.addWidget(browse_input_btn)
        layout.addLayout(input_layout)

        populate_btn = QPushButton("List Hives from Directory")
        small_button_style = self.get_button_style(bg_color=COLOR_DARK, text_color="white", hover_color=COLOR_ORANGE)
        small_button_style = small_button_style.replace("padding: 18px 64px;", "padding: 8px 12px;").replace("font-size: 22px;", "font-size: 14px;")
        populate_btn.setStyleSheet(small_button_style)
        populate_btn.setFixedHeight(44)
        populate_btn.clicked.connect(self.populate_hives_for_analysis)
        layout.addWidget(populate_btn, alignment=Qt.AlignLeft)
        
        layout.addSpacing(10)
        layout.addWidget(QLabel("Select Hives to Analyze:"))
        self.analyze_hive_list = QListWidget()
        self.analyze_hive_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        self.analyze_hive_list.setMaximumHeight(150)
        self.analyze_hive_list.setStyleSheet(f"border: 1px solid {COLOR_DARK}; border-radius: 5px; padding: 5px;")
        layout.addWidget(self.analyze_hive_list)

        layout.addSpacing(15)
        analyze_btn = self.create_styled_button("Analyze Selected Hives", self.analyze_hives)
        layout.addWidget(analyze_btn, alignment=Qt.AlignCenter)
        return group

    def create_compare_hives_group(self):
        group = QGroupBox("3. Compare Registry Hives")
        group.setFont(QFont("Cascadia Mono", 12, QFont.Weight.Bold))
        group.setStyleSheet(self._get_group_box_style())
        layout = QVBoxLayout(group)
        layout.setSpacing(10)
        
        layout.addWidget(QLabel("First Hive:"))
        hive1_layout = QHBoxLayout()
        self.hive1_input = self.create_styled_input("Path to first hive file")
        browse1_btn = self._create_small_browse_button(lambda: self.browse_file(self.hive1_input))
        hive1_layout.addWidget(self.hive1_input)
        hive1_layout.addWidget(browse1_btn)
        layout.addLayout(hive1_layout)

        layout.addSpacing(10)
        layout.addWidget(QLabel("Second Hive:"))
        hive2_layout = QHBoxLayout()
        self.hive2_input = self.create_styled_input("Path to second hive file")
        browse2_btn = self._create_small_browse_button(lambda: self.browse_file(self.hive2_input))
        hive2_layout.addWidget(self.hive2_input)
        hive2_layout.addWidget(browse2_btn)
        layout.addLayout(hive2_layout)
        
        layout.addSpacing(10)
        layout.addWidget(QLabel("Output Directory for Report:"))
        output_layout = QHBoxLayout()
        self.compare_output_dir = self.create_styled_input()
        browse3_btn = self._create_small_browse_button(lambda: self.browse_directory(self.compare_output_dir))
        output_layout.addWidget(self.compare_output_dir)
        output_layout.addWidget(browse3_btn)
        layout.addLayout(output_layout)

        layout.addSpacing(15)
        compare_btn = self.create_styled_button("Compare Hives", self.compare_hives)
        layout.addWidget(compare_btn, alignment=Qt.AlignCenter)
        
        return group

    def create_apply_logs_group(self):
        group = QGroupBox("4. Apply Transaction Logs")
        group.setFont(QFont("Cascadia Mono", 12, QFont.Weight.Bold))
        group.setStyleSheet(self._get_group_box_style())
        layout = QVBoxLayout(group)
        layout.setSpacing(10)

        layout.addWidget(QLabel("Hive File:"))
        hive_layout = QHBoxLayout()
        self.logs_hive_input = self.create_styled_input("Path to hive file (e.g., SYSTEM, NTUSER.DAT)")
        browse1_btn = self._create_small_browse_button(lambda: self.browse_file(self.logs_hive_input))
        hive_layout.addWidget(self.logs_hive_input)
        hive_layout.addWidget(browse1_btn)
        layout.addLayout(hive_layout)
        
        layout.addSpacing(10)
        layout.addWidget(QLabel("Output Directory for Recovered Hive:"))
        output_layout = QHBoxLayout()
        self.logs_output_dir = self.create_styled_input()
        browse2_btn = self._create_small_browse_button(lambda: self.browse_directory(self.logs_output_dir))
        output_layout.addWidget(self.logs_output_dir)
        output_layout.addWidget(browse2_btn)
        layout.addLayout(output_layout)

        layout.addSpacing(15)
        apply_btn = self.create_styled_button("Apply Transaction Logs", self.apply_transaction_logs)
        layout.addWidget(apply_btn, alignment=Qt.AlignCenter)
        
        return group

    def create_parse_header_group(self):
        group = QGroupBox("5. Parse Hive Header")
        group.setFont(QFont("Cascadia Mono", 12, QFont.Weight.Bold))
        group.setStyleSheet(self._get_group_box_style())
        layout = QVBoxLayout(group)
        layout.setSpacing(10)
        
        layout.addWidget(QLabel("Hive File:"))
        hive_layout = QHBoxLayout()
        self.header_hive_input = self.create_styled_input("Path to hive file to parse")
        browse_btn = self._create_small_browse_button(lambda: self.browse_file(self.header_hive_input))
        hive_layout.addWidget(self.header_hive_input)
        hive_layout.addWidget(browse_btn)
        layout.addLayout(hive_layout)

        layout.addSpacing(15)
        parse_btn = self.create_styled_button("Parse Hive Header", self.parse_hive_header)
        layout.addWidget(parse_btn, alignment=Qt.AlignCenter)
        
        return group
    
    def populate_hives_for_analysis(self):
        """Lists hive files from the selected input directory."""
        input_dir = self.analyze_input_dir.text()
        if not os.path.isdir(input_dir):
            QMessageBox.warning(self, "Invalid Directory", "Please select a valid directory first.")
            return
        self.analyze_hive_list.clear()
        try:
            for item in os.listdir(input_dir):
                if os.path.isfile(os.path.join(input_dir, item)):
                    self.analyze_hive_list.addItem(QListWidgetItem(item))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not read directory: {e}")

    def acquire_hives(self):
        """Handles the logic for acquiring selected hives."""
        selected_items = self.hive_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Missing Information", "Please select at least one hive to acquire.")
            return
        output_dir = self.acquire_output_dir_input.text()
        if not output_dir:
            QMessageBox.warning(self, "Missing Information", "Please specify an output directory.")
            return
        selected_hives = [item.text() for item in selected_items]
        username = self.username_input.text()
        self.start_registry_operation("acquire_registry_hives", {
            'output_dir': output_dir,
            'selected_hives': selected_hives,
            'username': username
        })

    def analyze_hives(self):
        """Handles the logic for analyzing selected hives."""
        selected_items = self.analyze_hive_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Missing Information", "Please select at least one hive to analyze.")
            return
        selected_hives = [item.text() for item in selected_items]
        analysis_dir = os.path.join(self.selected_case_path, "registry_analysis", "analysis_results")
        self.start_registry_operation("analyze_registry_hive", {
            'input_dir': self.analyze_input_dir.text(),
            'analysis_dir': analysis_dir,
            'selected_hives': selected_hives
        })

    def compare_hives(self):
        """Handles the logic for comparing two hives."""
        hive1 = self.hive1_input.text()
        hive2 = self.hive2_input.text()
        output_dir = self.compare_output_dir.text()
        if not all([hive1, hive2, output_dir]):
            QMessageBox.warning(self, "Missing Information", "Please provide paths for both hives and an output directory.")
            return
        self.start_registry_operation("compare_registry_hives", {
            'hive1_path': hive1,
            'hive2_path': hive2,
            'output_dir': output_dir
        })

    def apply_transaction_logs(self):
        """Handles the logic for applying transaction logs."""
        hive_path = self.logs_hive_input.text()
        output_dir = self.logs_output_dir.text()
        if not all([hive_path, output_dir]):
            QMessageBox.warning(self, "Missing Information", "Please provide the hive path and an output directory.")
            return
        self.start_registry_operation("apply_transaction_logs", {
            'hive_path': hive_path,
            'output_dir': output_dir
        })

    def parse_hive_header(self):
        """Handles the logic for parsing a hive header."""
        hive_path = self.header_hive_input.text()
        if not hive_path:
            QMessageBox.warning(self, "Missing Information", "Please provide the hive path.")
            return
        self.start_registry_operation("parse_hive_header", {'hive_path': hive_path})

    def browse_directory(self, input_field):
        """Opens a dialog to select a directory and sets the path to the input field."""
        directory = QFileDialog.getExistingDirectory(self, "Select Directory")
        if directory:
            input_field.setText(directory)

    def browse_file(self, input_field):
        """Opens a dialog to select a file and sets the path to the input field."""
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Hive File", "", "All Files (*)")
        if file_path:
            input_field.setText(file_path)

    def start_registry_operation(self, operation, kwargs):
        if self.registry_worker_thread and self.registry_worker_thread.isRunning():
            QMessageBox.warning(self, "In Progress", "Another registry operation is in progress.")
            return
        
        self.registry_progress_text.clear()
        self.registry_worker_thread = RegistryWorker(self.registry_analyzer, operation, **kwargs)
        self.registry_worker_thread.progress_updated.connect(self.update_registry_progress)
        self.registry_worker_thread.operation_completed.connect(self.handle_registry_operation_completed)
        self.registry_worker_thread.header_output.connect(self.display_header_output)
        self.registry_worker_thread.start()

    def update_registry_progress(self, message):
        self.registry_progress_text.append(message)

    def handle_registry_operation_completed(self, operation, success, message):
        status = "SUCCESS" if success else "FAILED"
        self.registry_progress_text.append(f"--- [{datetime.now().strftime('%H:%M:%S')}] {operation.replace('_', ' ').title()} {status} ---")
        if not success:
             self.registry_progress_text.append(f"Error: {message}\n")
        else:
             self.registry_progress_text.append(f"Details: {message}\n")
        
        # No popup for every operation, progress text is enough
        # QMessageBox.information(self, f"Operation {status}", message)

    def display_header_output(self, output):
        """Display header parsing output in a formatted way"""
        self.registry_progress_text.append("=" * 60)
        self.registry_progress_text.append("REGISTRY HIVE HEADER ANALYSIS")
        self.registry_progress_text.append("=" * 60)
        self.registry_progress_text.append(output)
        self.registry_progress_text.append("=" * 60)
        self.registry_progress_text.append("")  # Add empty line for spacing

if __name__ == '__main__':
    import sys
    print("This module is part of the Anubis Forensics GUI application.")
    print("Please run the main application using: python main.py")
    sys.exit(1) 
